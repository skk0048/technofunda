"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   RS REBALANCING TRACKER  v2.0  — Cap-Tier Separated Sleeves              ║
║                                                                            ║
║   WHAT'S NEW IN v2.0:                                                      ║
║   ✅ Interactive menu  — just run the .py and pick 1-9. No CLI args.      ║
║   ✅ Past-date analysis — backtest any date; type YYYY-MM-DD in menu      ║
║   ✅ USA expanded      — US_A (Top 50) · US_B (51-200) · US_C (201-500)  ║
║   ✅ US_ETF sleeve     — 37 World + Commodity + Defensive ETFs            ║
║   ✅ Snapshot compare  — auto-saves JSON after every run; next run shows  ║
║                          NEW / REMOVED / ADJUST-QTY vs last snapshot      ║
║                                                                            ║
║   INDIA                                                                    ║
║   Sleeve A    — Large Cap   (Nifty 1-50)       Monthly   · Long-Heavy RS  ║
║   Sleeve B    — Mid-Large   (Nifty 51-200)     Fortnight · Balanced RS    ║
║   Sleeve C    — Small-Mid   (Nifty 201-500)    Weekly    · Short-Heavy RS ║
║   Sleeve D    — Liquid Bees (bear buffer)       As-needed                 ║
║   USA                                                                      ║
║   Sleeve US_A — Mega Cap  (S&P Top 50)          Monthly                   ║
║   Sleeve US_B — Large Cap (S&P 51-200)          Fortnight                 ║
║   Sleeve US_C — Mid Cap   (S&P 201-500)         Weekly                    ║
║   Sleeve US_ETF — Global ETFs (World+Comm+Def)  Monthly                   ║
║                                                                            ║
║   NOTE: us_sp500list.csv must be sorted by market cap (largest first)    ║
║         so row ranges correctly separate Mega / Large / Mid cap tiers.   ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import os, sys, json, glob, warnings, argparse
from datetime import timedelta, date
from pathlib import Path

import pandas as pd
import numpy as np
import yfinance as yf
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════════════════
#  ❶  CONFIGURATION  (edit these paths to match your machine)
# ══════════════════════════════════════════════════════════════════════════════

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))

# Portable path resolution (#10): prefer the hardcoded machine paths when they
# exist (keeps the current setup working), else fall back to locations relative
# to this script so the project can be dropped into any folder.
_HC_BASE  = r"C:\Users\sudhi\Documents\Trading\SectorRotation\SectorRotationBacktest"
_HC_INDEX = r"C:\Users\sudhi\Documents\Trading\SectorRotation\IndexData"

BASE_DIR  = _HC_BASE if os.path.isdir(_HC_BASE) else SCRIPT_DIR
INDEX_DIR = next(
    (d for d in (os.path.join(SCRIPT_DIR, "SupportFiles", "IndexData"),
                 os.path.join(SCRIPT_DIR, "IndexData"),
                 _HC_INDEX) if os.path.isdir(d)),
    _HC_INDEX,
)
CACHE_DIR    = r"C:\Users\sudhi\AppData\Local\StockPriceCache"
HOLDINGS_CSV = os.path.join(BASE_DIR, "my_holdings.csv")
OUTPUT_DIR   = BASE_DIR
SNAPSHOT_DIR = os.path.join(OUTPUT_DIR, "snapshots")   # ← NEW: auto-created

# --- ATR Position Sizing ---
ATR_PERIOD     = 14
VOL_MIN_WEIGHT = 0.25
VOL_MAX_WEIGHT = 2.50
USE_VOL_SIZING = True

# --- Global Analysis Date (set by menu or --date flag) ---
# Change this via the interactive menu or --date YYYY-MM-DD for backtesting
ANALYSIS_DATE = date.today()

# ══════════════════════════════════════════════════════════════════════════════
#  ❷  US ETF HARDCODED UNIVERSE  (no CSV needed — edit list as you like)
# ══════════════════════════════════════════════════════════════════════════════

US_ETF_UNIVERSE = [
    # ── Broad / Regional ETFs ────────────────────────────────────────────────
    {"Symbol": "VEA",  "Company Name": "Vanguard Dev Markets ETF",        "Industry": "World ETF",     "Sector": "World"},
    {"Symbol": "VWO",  "Company Name": "Vanguard Emerging Markets ETF",   "Industry": "EM ETF",        "Sector": "World"},
    {"Symbol": "EEM",  "Company Name": "iShares MSCI Emerging Markets",   "Industry": "EM ETF",        "Sector": "World"},
    {"Symbol": "EFA",  "Company Name": "iShares MSCI EAFE",               "Industry": "World ETF",     "Sector": "World"},
    {"Symbol": "VT",   "Company Name": "Vanguard Total World Stock",      "Industry": "World ETF",     "Sector": "World"},
    {"Symbol": "ACWI", "Company Name": "iShares MSCI ACWI",               "Industry": "World ETF",     "Sector": "World"},
    {"Symbol": "VEU",  "Company Name": "Vanguard FTSE All-World ex-US",   "Industry": "World ETF",     "Sector": "World"},
    {"Symbol": "VGK",  "Company Name": "Vanguard FTSE Europe ETF",        "Industry": "World ETF",     "Sector": "World"},
    {"Symbol": "VPL",  "Company Name": "Vanguard FTSE Pacific ETF",       "Industry": "World ETF",     "Sector": "World"},
    {"Symbol": "FM",   "Company Name": "iShares MSCI Frontier Markets",   "Industry": "EM ETF",        "Sector": "World"},
    # ── Americas ─────────────────────────────────────────────────────────────
    {"Symbol": "EWC",  "Company Name": "iShares Canada ETF",              "Industry": "Country ETF",   "Sector": "Americas"},
    {"Symbol": "EWZ",  "Company Name": "iShares Brazil ETF",              "Industry": "Country ETF",   "Sector": "Americas"},
    {"Symbol": "EWW",  "Company Name": "iShares Mexico ETF",              "Industry": "Country ETF",   "Sector": "Americas"},
    {"Symbol": "ECH",  "Company Name": "iShares Chile ETF",               "Industry": "Country ETF",   "Sector": "Americas"},
    {"Symbol": "GXG",  "Company Name": "Global X Colombia ETF",           "Industry": "Country ETF",   "Sector": "Americas"},
    {"Symbol": "EPU",  "Company Name": "iShares Peru ETF",                "Industry": "Country ETF",   "Sector": "Americas"},
    {"Symbol": "ARGT", "Company Name": "Global X Argentina ETF",          "Industry": "Country ETF",   "Sector": "Americas"},
    # ── Europe ───────────────────────────────────────────────────────────────
    {"Symbol": "EWU",  "Company Name": "iShares UK ETF",                  "Industry": "Country ETF",   "Sector": "Europe"},
    {"Symbol": "EWG",  "Company Name": "iShares Germany ETF",             "Industry": "Country ETF",   "Sector": "Europe"},
    {"Symbol": "EWQ",  "Company Name": "iShares France ETF",              "Industry": "Country ETF",   "Sector": "Europe"},
    {"Symbol": "EWI",  "Company Name": "iShares Italy ETF",               "Industry": "Country ETF",   "Sector": "Europe"},
    {"Symbol": "EWP",  "Company Name": "iShares Spain ETF",               "Industry": "Country ETF",   "Sector": "Europe"},
    {"Symbol": "EWN",  "Company Name": "iShares Netherlands ETF",         "Industry": "Country ETF",   "Sector": "Europe"},
    {"Symbol": "EWL",  "Company Name": "iShares Switzerland ETF",         "Industry": "Country ETF",   "Sector": "Europe"},
    {"Symbol": "EWD",  "Company Name": "iShares Sweden ETF",              "Industry": "Country ETF",   "Sector": "Europe"},
    {"Symbol": "EPOL", "Company Name": "iShares Poland ETF",              "Industry": "Country ETF",   "Sector": "Europe"},
    {"Symbol": "TUR",  "Company Name": "iShares Turkey ETF",              "Industry": "Country ETF",   "Sector": "Europe"},
    {"Symbol": "GREK", "Company Name": "Global X Greece ETF",             "Industry": "Country ETF",   "Sector": "Europe"},
    # ── Middle East / Africa ─────────────────────────────────────────────────
    {"Symbol": "KSA",  "Company Name": "iShares Saudi Arabia ETF",        "Industry": "Country ETF",   "Sector": "Mid East"},
    {"Symbol": "UAE",  "Company Name": "iShares UAE ETF",                  "Industry": "Country ETF",   "Sector": "Mid East"},
    {"Symbol": "EIS",  "Company Name": "iShares Israel ETF",              "Industry": "Country ETF",   "Sector": "Mid East"},
    {"Symbol": "EZA",  "Company Name": "iShares South Africa ETF",        "Industry": "Country ETF",   "Sector": "Africa"},
    {"Symbol": "EGPT", "Company Name": "VanEck Egypt ETF",                "Industry": "Country ETF",   "Sector": "Africa"},
    {"Symbol": "NGE",  "Company Name": "Global X Nigeria ETF",            "Industry": "Country ETF",   "Sector": "Africa"},
    # ── Asia-Pacific ─────────────────────────────────────────────────────────
    {"Symbol": "MCHI", "Company Name": "iShares MSCI China ETF",          "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "FXI",  "Company Name": "iShares China Large Cap ETF",     "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "INDA", "Company Name": "iShares India ETF",               "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "EWJ",  "Company Name": "iShares Japan ETF",               "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "EWY",  "Company Name": "iShares South Korea ETF",         "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "EWT",  "Company Name": "iShares Taiwan ETF",              "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "EWH",  "Company Name": "iShares Hong Kong ETF",           "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "EWS",  "Company Name": "iShares Singapore ETF",           "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "EWA",  "Company Name": "iShares Australia ETF",           "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "VNM",  "Company Name": "VanEck Vietnam ETF",              "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "EIDO", "Company Name": "iShares Indonesia ETF",           "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "THD",  "Company Name": "iShares Thailand ETF",            "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "EWM",  "Company Name": "iShares Malaysia ETF",            "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "EPHE", "Company Name": "iShares Philippines ETF",         "Industry": "Country ETF",   "Sector": "Asia"},
    {"Symbol": "PAK",  "Company Name": "Global X Pakistan ETF",           "Industry": "Country ETF",   "Sector": "Asia"},
    # ── Precious Metals ───────────────────────────────────────────────────────
    {"Symbol": "GLD",  "Company Name": "SPDR Gold Shares",                "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "IAU",  "Company Name": "iShares Gold Trust",              "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "SLV",  "Company Name": "iShares Silver Trust",            "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "PPLT", "Company Name": "Aberdeen Platinum ETF",           "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "PALL", "Company Name": "Aberdeen Palladium ETF",          "Industry": "Commodity ETF", "Sector": "Commodity"},
    # ── Energy Commodities ────────────────────────────────────────────────────
    {"Symbol": "USO",  "Company Name": "US Oil Fund (WTI Crude)",         "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "BNO",  "Company Name": "US Brent Oil Fund",               "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "UNG",  "Company Name": "US Natural Gas Fund",             "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "UGA",  "Company Name": "US Gasoline Fund",                "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "URA",  "Company Name": "Global X Uranium ETF",            "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "ARCH", "Company Name": "Arch Resources (Coal proxy)",     "Industry": "Commodity ETF", "Sector": "Commodity"},
    # ── Base / Industrial Metals ──────────────────────────────────────────────
    {"Symbol": "CPER", "Company Name": "US Copper Index Fund",            "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "PICK", "Company Name": "iShares MSCI Global Metals ETF",  "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "SLX",  "Company Name": "VanEck Steel ETF",                "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "LIT",  "Company Name": "Global X Lithium ETF",            "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "REMX", "Company Name": "VanEck Rare Earth ETF",           "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "DBB",  "Company Name": "Invesco DB Base Metals Fund",     "Industry": "Commodity ETF", "Sector": "Commodity"},
    # ── Agriculture ───────────────────────────────────────────────────────────
    {"Symbol": "CORN", "Company Name": "Teucrium Corn Fund",              "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "WEAT", "Company Name": "Teucrium Wheat Fund",             "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "SOYB", "Company Name": "Teucrium Soybean Fund",           "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "CANE", "Company Name": "Teucrium Sugar Fund",             "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "JO",   "Company Name": "iPath Coffee ETN",                "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "BAL",  "Company Name": "iPath Cotton ETN",                "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "CUT",  "Company Name": "Invesco MSCI Global Timber ETF",  "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "DBA",  "Company Name": "Invesco DB Agriculture",          "Industry": "Commodity ETF", "Sector": "Commodity"},
    # ── Broad Commodity ───────────────────────────────────────────────────────
    {"Symbol": "PDBC", "Company Name": "Invesco Optimum Yield Commodity", "Industry": "Commodity ETF", "Sector": "Commodity"},
    {"Symbol": "DBE",  "Company Name": "Invesco DB Energy Fund",          "Industry": "Commodity ETF", "Sector": "Commodity"},
    # ── Defensive / Bonds ─────────────────────────────────────────────────────
    {"Symbol": "TLT",  "Company Name": "iShares 20yr Treasury Bond",      "Industry": "Bond ETF",      "Sector": "Defensive"},
    {"Symbol": "IEF",  "Company Name": "iShares 7-10yr Treasury Bond",    "Industry": "Bond ETF",      "Sector": "Defensive"},
    {"Symbol": "AGG",  "Company Name": "iShares US Aggregate Bond",       "Industry": "Bond ETF",      "Sector": "Defensive"},
    {"Symbol": "TIP",  "Company Name": "iShares TIPS Bond ETF",           "Industry": "Bond ETF",      "Sector": "Defensive"},
    {"Symbol": "SHY",  "Company Name": "iShares 1-3yr Treasury Bond",     "Industry": "Bond ETF",      "Sector": "Defensive"},
    {"Symbol": "LQD",  "Company Name": "iShares Investment Grade Corp",   "Industry": "Bond ETF",      "Sector": "Defensive"},
    {"Symbol": "HYG",  "Company Name": "iShares High Yield Corp Bond",    "Industry": "Bond ETF",      "Sector": "Defensive"},
    # ── Defensive / Real Assets ───────────────────────────────────────────────
    {"Symbol": "VNQ",  "Company Name": "Vanguard Real Estate REIT",       "Industry": "REIT ETF",      "Sector": "Defensive"},
    {"Symbol": "XLU",  "Company Name": "Utilities Select SPDR",           "Industry": "Sector ETF",    "Sector": "Defensive"},
    {"Symbol": "XLP",  "Company Name": "Consumer Staples SPDR",           "Industry": "Sector ETF",    "Sector": "Defensive"},
    # ── US Sector ETFs ────────────────────────────────────────────────────────
    {"Symbol": "XLK",  "Company Name": "Technology Select SPDR",          "Industry": "Sector ETF",    "Sector": "Tech"},
    {"Symbol": "XLF",  "Company Name": "Financials Select SPDR",          "Industry": "Sector ETF",    "Sector": "Financials"},
    {"Symbol": "XLV",  "Company Name": "Healthcare Select SPDR",          "Industry": "Sector ETF",    "Sector": "Healthcare"},
    {"Symbol": "XLY",  "Company Name": "Consumer Disc SPDR",              "Industry": "Sector ETF",    "Sector": "ConsumerDisc"},
    {"Symbol": "XLI",  "Company Name": "Industrials Select SPDR",         "Industry": "Sector ETF",    "Sector": "Industrials"},
    {"Symbol": "XLE",  "Company Name": "Energy Select SPDR",              "Industry": "Sector ETF",    "Sector": "Energy"},
    {"Symbol": "XLB",  "Company Name": "Materials Select SPDR",           "Industry": "Sector ETF",    "Sector": "Materials"},
    {"Symbol": "XLRE", "Company Name": "Real Estate Select SPDR",         "Industry": "Sector ETF",    "Sector": "RealEstate"},
    {"Symbol": "XLC",  "Company Name": "Communication Services SPDR",     "Industry": "Sector ETF",    "Sector": "CommServices"},
]

# ══════════════════════════════════════════════════════════════════════════════
#  ❸  SLEEVES  (India A/B/C/D unchanged · USA expanded to 4 sleeves)
# ══════════════════════════════════════════════════════════════════════════════
#
#  universe_row_range: (start, end) — row slice of the CSV (0-indexed).
#    Used for US cap-tier splitting WITHOUT needing separate CSV files.
#    REQUIREMENT: us_sp500list.csv must be sorted by market cap, largest first.
#
SLEEVES = {
    # ── INDIA ─────────────────────────────────────────────────────────────────
    "A": {
        "name":              "Core",
        "cap_tier":          "Large Cap  (Nifty 1-50)",
        "market":            "INDIA",
        "universe":          "ind_nifty50list.csv",
        "exclude_universe":  None,
        "universe_row_range": None,
        "top_n":             10,
        "stop_loss":         0.15,
        "take_profit":       0.0,
        "rebalance":         "monthly",
        "rs_weights":        {22: 0.40, 55: 0.30, 120: 0.20, 252: 0.10},
        "capital":           3_500_000,
        "description":       "Large Cap Momentum | Nifty 50 | Monthly | Long-Heavy RS",
    },
    "B": {
        "name":              "Growth",
        "cap_tier":          "Mid-Large Cap  (Nifty 51-200)",
        "market":            "INDIA",
        "universe":          "ind_nifty200list.csv",
        "exclude_universe":  "ind_nifty50list.csv",
        "universe_row_range": None,
        "top_n":             15,
        "stop_loss":         0.20,
        "take_profit":       0.60,
        "rebalance":         "fortnightly",
        "rs_weights":        {22: 0.50, 55: 0.25, 120: 0.20, 252: 0.05},
        "capital":           3_000_000,
        "description":       "Mid-Large Cap Momentum | Nifty 51-200 | Fortnightly | Balanced RS",
    },
    "C": {
        "name":              "Aggressive",
        "cap_tier":          "Small-Mid Cap  (Nifty 201-500)",
        "market":            "INDIA",
        "universe":          "ind_nifty500list.csv",
        "exclude_universe":  "ind_nifty200list.csv",
        "universe_row_range": None,
        "top_n":             15,
        "stop_loss":         0.25,
        "take_profit":       0.50,
        "rebalance":         "weekly",
        "rs_weights":        {22: 0.60, 55: 0.25, 120: 0.10, 252: 0.05},
        "capital":           2_000_000,
        "description":       "Small-Mid Cap Momentum | Nifty 201-500 | Weekly | Short-Heavy RS",
    },
    "D": {
        "name":              "Liquid",
        "cap_tier":          "Cash / Liquid Bees",
        "market":            "INDIA",
        "universe":          None,
        "exclude_universe":  None,
        "universe_row_range": None,
        "top_n":             0,
        "stop_loss":         0.0,
        "take_profit":       0.0,
        "rebalance":         "as_needed",
        "rs_weights":        {},
        "capital":           1_500_000,
        "description":       "Liquid Bees / FD | Bear buffer + dry powder | No RS scoring",
    },
    # ── USA — 3 cap tiers matching India structure ────────────────────────────
    # NOTE: us_sp500list.csv must be sorted by market cap descending (largest first).
    #       Row 0-49  = Top 50 mega caps  → US_A
    #       Row 50-199 = Large 51-200     → US_B
    #       Row 200-499 = Mid 201-500     → US_C
    "US_A": {
        "name":              "US Mega",
        "cap_tier":          "US Mega Cap  (S&P Top 50)",
        "market":            "US",
        "universe":          "us_sp500list.csv",
        "exclude_universe":  None,
        "universe_row_range": (0, 50),           # first 50 rows = mega cap
        "top_n":             15,
        "stop_loss":         0.10,
        "take_profit":       0.40,
        "rebalance":         "monthly",
        "rs_weights":        {22: 0.30, 55: 0.30, 120: 0.25, 252: 0.15},  # Long-heavy
        "capital":           0,
        "description":       "US Mega Cap Momentum | S&P Top 50 | Monthly | Long-Heavy RS",
    },
    "US_B": {
        "name":              "US Large",
        "cap_tier":          "US Large Cap  (S&P 51-200)",
        "market":            "US",
        "universe":          "us_sp500list.csv",
        "exclude_universe":  None,
        "universe_row_range": (50, 200),         # rows 50-199 = large cap
        "top_n":             20,
        "stop_loss":         0.12,
        "take_profit":       0.45,
        "rebalance":         "fortnightly",
        "rs_weights":        {22: 0.40, 55: 0.30, 120: 0.20, 252: 0.10},  # Balanced
        "capital":           0,
        "description":       "US Large Cap Momentum | S&P 51-200 | Fortnightly | Balanced RS",
    },
    "US_C": {
        "name":              "US Mid",
        "cap_tier":          "US Mid Cap  (S&P 201-500)",
        "market":            "US",
        "universe":          "us_sp500list.csv",
        "exclude_universe":  None,
        "universe_row_range": (200, 500),        # rows 200-499 = mid cap
        "top_n":             20,
        "stop_loss":         0.15,
        "take_profit":       0.50,
        "rebalance":         "weekly",
        "rs_weights":        {22: 0.50, 55: 0.30, 120: 0.15, 252: 0.05},  # Short-heavy
        "capital":           0,
        "description":       "US Mid Cap Momentum | S&P 201-500 | Weekly | Short-Heavy RS",
    },
    "US_ETF": {
        "name":              "Global ETF",
        "cap_tier":          "World + Commodity + Defensive ETFs",
        "market":            "US",
        "universe":          "__ETF_HARDCODED__",  # special marker — no CSV needed
        "exclude_universe":  None,
        "universe_row_range": None,
        "top_n":             10,
        "stop_loss":         0.08,
        "take_profit":       0.30,
        "rebalance":         "monthly",
        "rs_weights":        {22: 0.40, 55: 0.35, 120: 0.20, 252: 0.05},
        "capital":           0,
        "description":       "Global ETF Rotation | World + Commodities + Defensive | Monthly",
    },
}

# ── Market configs ────────────────────────────────────────────────────────────
INDIA_INDEX   = "^NSEI"
INDIA_SECTORS = {
    "Automobile": "^CNXAUTO", "IT": "^CNXIT",
    "Banking": "^NSEBANK",    "Pharma": "^CNXPHARMA",
    "FMCG": "^CNXFMCG",       "Metal": "^CNXMETAL",
    "OilGas": "^CNXENERGY",   "Finance": "^CNXFIN",
    "Realty": "^CNXREALTY",   "Infra": "^CNXINFRA",
}
US_INDEX   = "SPY"
US_SECTORS = {
    "Technology": "XLK", "Financials": "XLF", "Healthcare": "XLV",
    "ConsumerDisc": "XLY", "Industrials": "XLI", "Energy": "XLE",
}
INDIA_IND_MAP = {
    "Automobile and Auto Components": "Automobile", "Information Technology": "IT",
    "Financial Services": "Finance", "Healthcare": "Healthcare",
    "Fast Moving Consumer Goods": "FMCG", "Metals & Mining": "Metal",
    "Oil, Gas & Consumable Fuels": "OilGas", "Oil Gas & Consumable Fuels": "OilGas",
    "Realty": "Realty", "Capital Goods": "Infra", "Construction": "Infra",
    "Construction Materials": "Cement", "Media, Entertainment & Publication": "Media",
    "Chemicals": "Chemicals", "Consumer Durables": "ConsumerDur",
    "Consumer Services": "ConsumerDur", "Power": "OilGas", "Telecommunication": "IT",
}
US_IND_MAP = {
    "Information Technology": "Technology", "Semiconductors": "Technology",
    "Financials": "Financials", "Health Care": "Healthcare",
    "Consumer Discretionary": "ConsumerDisc", "Industrials": "Industrials",
    "Energy": "Energy", "Materials": "Materials",
}

REGIME_EMA_FAST = 100
REGIME_EMA_SLOW = 200
REGIME_EXPOSURE = {"BULL": 1.00, "CAUTION": 0.50, "BEAR": 0.25}

STRICT_PEER_FILTER = True
PEER_PERIOD        = 55

# ── RS Decision Period ────────────────────────────────────────────────────────
# This is the primary RS look-back period used for the PEER filter (stock vs
# sector vs index comparison).  Change this single value to switch between
# short-term (22) and medium-term (55) signals across ALL sleeves at once.
#   22  = faster, reacts in ~1 month  (recommended for current use)
#   55  = slower, reacts in ~2.5 months
RS_DECISION_PERIOD = 22          # ← CHANGE THIS to switch 22 / 55

SECTOR_CAP_PCT     = 0.25
MIN_RS_SCORE       = 0.0
VOLUME_LOOKBACK    = 14
MIN_TURNOVER_CR    = 5.0
MIN_TURNOVER_USD   = 5.0
LOOKBACK_DAYS      = 380

# ── Action List: demo capital per sleeve for Est Qty calculation ──────────────
# Columns J & K in the BUY/HOLD table ("Capital" and "Est Qty") are computed
# against this fixed notional value.  Change to any amount you like.
SLEEVE_DEMO_CAPITAL = 10_000     # ← 10 K per sleeve (USD or INR, same logic)

# ══════════════════════════════════════════════════════════════════════════════
#  ❹  CACHE LAYER  (unchanged from v1.2)
# ══════════════════════════════════════════════════════════════════════════════

class SimpleCache:
    def __init__(self, cache_dir):
        self.base = Path(cache_dir)
        for sub in ["_meta", "close", "low", "high", "volume"]:
            (self.base / sub).mkdir(parents=True, exist_ok=True)

    def _key(self, s): return s.replace("^", "IDX_").replace(".", "_")
    def _path(self, sub, s): return self.base / sub / f"{self._key(s)}.parquet"
    def _meta(self, s): return self.base / "_meta" / f"{self._key(s)}.json"

    def has(self, sym, start, end):
        try:
            m = json.loads(self._meta(sym).read_text())
            return (m.get("version", "0") >= "2.1"
                    and m.get("start", "9") <= start
                    and m.get("end", "0") >= end
                    and m.get("n_rows", 0) > 50)
        except:
            return False

    def load(self, sym, sub="close"):
        try:
            df  = pd.read_parquet(self._path(sub, sym))
            s   = df.iloc[:, 0].dropna().sort_index()
            idx = s.index
            if hasattr(idx, "tz") and idx.tz:
                idx = idx.tz_localize(None)
            idx = idx.normalize()
            s   = pd.Series(s.values, index=idx)
            return s[~s.index.duplicated(keep="last")].astype(float)
        except:
            return pd.Series(dtype=float)

    def save(self, sym, c, l, h, v=None):
        if c.empty: return
        for sub, s in [("close", c), ("low", l), ("high", h)]:
            if not s.empty:
                pd.DataFrame({sym: s}).to_parquet(self._path(sub, sym))
        if v is not None and not v.empty:
            pd.DataFrame({sym: v}).to_parquet(self._path("volume", sym))
        self._meta(sym).write_text(json.dumps({
            "version":      "2.1", "symbol": sym,
            "start":        str(c.index.min().date()),
            "end":          str(c.index.max().date()),
            "n_rows":       len(c),
            "last_updated": str(date.today()),
        }))

    def fetch(self, symbols, start, end):
        need = [s for s in symbols if not self.has(s, start, end)]
        if need:
            print(f"    Downloading {len(need)}/{len(symbols)} symbols …")
            for i in range(0, len(need), 100):
                batch = need[i:i+100]
                try:
                    raw = yf.download(batch, start=start, end=end,
                                      auto_adjust=True, progress=False, threads=True)
                    if isinstance(raw.columns, pd.MultiIndex):
                        cd, ld, hd, vd = raw["Close"], raw["Low"], raw["High"], raw["Volume"]
                    else:
                        cd = raw[["Close"]]; cd.columns = [batch[0]]
                        ld = raw[["Low"]];   ld.columns = [batch[0]]
                        hd = raw[["High"]];  hd.columns = [batch[0]]
                        vd = raw[["Volume"]]; vd.columns = [batch[0]]
                    for sym in batch:
                        if sym in cd.columns:
                            c = cd[sym].dropna()
                            if len(c) >= 50:
                                self.save(sym, c,
                                    ld[sym].dropna() if sym in ld.columns else pd.Series(dtype=float),
                                    hd[sym].dropna() if sym in hd.columns else pd.Series(dtype=float),
                                    vd[sym].dropna() if sym in vd.columns else pd.Series(dtype=float))
                except Exception as e:
                    print(f"    Batch error: {e}")
        else:
            print(f"    All {len(symbols)} in cache")

    def get_all(self, symbols, start, end):
        self.fetch(symbols, start, end)
        c, l, h, v = {}, {}, {}, {}
        for s in symbols:
            cl = self.load(s)
            if len(cl) >= 50:
                c[s] = cl
                l[s] = self.load(s, "low")
                h[s] = self.load(s, "high")
                v[s] = self.load(s, "volume")
        return pd.DataFrame(c), pd.DataFrame(l), pd.DataFrame(h), pd.DataFrame(v)


# ══════════════════════════════════════════════════════════════════════════════
#  ❺  UNIVERSE LOADER  — supports row_range for US cap-tiers + ETF hardcoded
# ══════════════════════════════════════════════════════════════════════════════

def _find_csv(fname):
    for d in [INDEX_DIR, BASE_DIR]:
        fp = os.path.join(d, fname)
        if os.path.exists(fp):
            return fp
    raise FileNotFoundError(f"File not found: {fname}\nLooked in: {INDEX_DIR}, {BASE_DIR}")


def load_universe(sleeve_cfg):
    """
    Load the stock / ETF universe for a sleeve.

    Handles three modes:
      1. Standard India CSV  (exclude_universe for cap-tier separation)
      2. US CSV with universe_row_range (0-indexed row slice, CSV sorted by mkt cap)
      3. __ETF_HARDCODED__  (no CSV — uses US_ETF_UNIVERSE list above)
    """
    if not sleeve_cfg.get("universe"):
        return pd.DataFrame()       # Sleeve D — Liquid, no stocks

    # ── Mode 3: hardcoded ETF list ─────────────────────────────────────────
    if sleeve_cfg["universe"] == "__ETF_HARDCODED__":
        df = pd.DataFrame(US_ETF_UNIVERSE).copy()
        df["Yahoo"] = df["Symbol"]  # ETFs trade on US exchanges — no suffix
        print(f"  Universe: ETF hardcoded list = {len(df)} ETFs  ← {sleeve_cfg['cap_tier']}")
        return df[["Company Name", "Symbol", "Yahoo", "Industry", "Sector"]].copy()

    market = sleeve_cfg["market"]
    imap   = INDIA_IND_MAP if market == "INDIA" else US_IND_MAP
    suffix = ".NS" if market == "INDIA" else ""

    df = pd.read_csv(_find_csv(sleeve_cfg["universe"]))
    df.columns   = df.columns.str.strip()
    df["Symbol"] = df["Symbol"].astype(str).str.strip()

    # Normalise Industry / Sector — only map if column exists
    if "Industry" in df.columns:
        df["Industry"] = df["Industry"].astype(str).str.strip()
        df["Sector"]   = df["Industry"].map(imap).fillna("Other")
    else:
        df["Industry"] = "Unknown"
        df["Sector"]   = "Other"

    df["Yahoo"] = df["Symbol"].apply(lambda s: s.replace(".", "-") + suffix)

    # ── Mode 2: row-range slice for US cap-tier separation ──────────────────
    row_range = sleeve_cfg.get("universe_row_range")
    if row_range:
        start_row, end_row = row_range
        n_before = len(df)
        df = df.iloc[start_row:end_row].reset_index(drop=True)
        print(f"  Universe: {sleeve_cfg['universe']} rows {start_row+1}–{end_row} "
              f"({n_before} total → {len(df)} stocks)  ← {sleeve_cfg['cap_tier']}")

    # ── Mode 1: India-style exclusion list ──────────────────────────────────
    else:
        excl_fname = sleeve_cfg.get("exclude_universe")
        if excl_fname:
            excl = pd.read_csv(_find_csv(excl_fname))
            excl.columns = excl.columns.str.strip()
            excl_syms  = set(excl["Symbol"].astype(str).str.strip())
            n_before   = len(df)
            df = df[~df["Symbol"].isin(excl_syms)].reset_index(drop=True)
            print(f"  Universe: {sleeve_cfg['universe']} ({n_before}) minus "
                  f"{excl_fname} ({len(excl_syms)}) = {len(df)} stocks"
                  f"  ← {sleeve_cfg['cap_tier']}")
        else:
            print(f"  Universe: {sleeve_cfg['universe']} = {len(df)} stocks"
                  f"  ← {sleeve_cfg['cap_tier']}")

    return df[["Company Name", "Symbol", "Yahoo", "Industry", "Sector"]].copy()


# ══════════════════════════════════════════════════════════════════════════════
#  ❻  OVERLAP DETECTION
# ══════════════════════════════════════════════════════════════════════════════

def check_universe_overlap(all_tops: dict):
    print("\n" + "─" * 60)
    print("  OVERLAP CHECK — stocks appearing in multiple sleeves")
    print("─" * 60)
    all_pairs  = [(a, b) for i, a in enumerate(all_tops) for b in list(all_tops)[i+1:]]
    any_overlap = False
    for s1, s2 in all_pairs:
        df1, df2 = all_tops[s1], all_tops[s2]
        if df1.empty or df2.empty: continue
        common = set(df1["Symbol"].tolist()) & set(df2["Symbol"].tolist())
        if common:
            any_overlap = True
            print(f"  ⚠ {s1} ∩ {s2}: {len(common)} overlap(s) — {', '.join(sorted(common))}")
        else:
            print(f"  ✅ {s1} ∩ {s2}: 0 overlap (as expected)")
    if not any_overlap:
        print("\n  ✅ PERFECT: Zero stocks shared across any sleeve pair.")
    else:
        print("\n  ⚠ Unexpected overlap — check exclude_universe / row_range settings.")
    print("─" * 60)


# ══════════════════════════════════════════════════════════════════════════════
#  ❼  REGIME DETECTION
# ══════════════════════════════════════════════════════════════════════════════

def get_regime(idx_series):
    if len(idx_series) < REGIME_EMA_SLOW + 10:
        return "BULL", 1.0, {}
    ema_slow = idx_series.ewm(span=REGIME_EMA_SLOW, adjust=False).mean()
    ema_fast = idx_series.ewm(span=REGIME_EMA_FAST, adjust=False).mean()
    now = float(idx_series.iloc[-1])
    es  = float(ema_slow.iloc[-1])
    ef  = float(ema_fast.iloc[-1])
    if now > es:   label, exp = "BULL",    REGIME_EXPOSURE["BULL"]
    elif now > ef: label, exp = "CAUTION", REGIME_EXPOSURE["CAUTION"]
    else:          label, exp = "BEAR",    REGIME_EXPOSURE["BEAR"]
    return label, exp, {"Index": round(now, 2), "EMA100": round(ef, 2), "EMA200": round(es, 2)}


# ══════════════════════════════════════════════════════════════════════════════
#  ❽  ATR WEIGHTS
# ══════════════════════════════════════════════════════════════════════════════

def calc_atr_weights(top_df, close_df, today):
    if top_df.empty: return {}, {}
    n = len(top_df)
    eq_w = 1.0 / n
    inv_vols, daily_stds = {}, {}
    for _, row in top_df.iterrows():
        sym = row["Yahoo"]
        try:
            prices = close_df[sym][close_df[sym].index <= today].tail(ATR_PERIOD + 5)
            std    = prices.pct_change().dropna().tail(ATR_PERIOD).std()
            if std > 0:
                inv_vols[sym]   = 1.0 / std
                daily_stds[sym] = round(std * 100, 3)
            else:
                inv_vols[sym]   = 1.0
                daily_stds[sym] = 0.0
        except Exception:
            inv_vols[sym]   = 1.0
            daily_stds[sym] = 0.0

    if not USE_VOL_SIZING:
        return {sym: eq_w for sym in inv_vols}, daily_stds

    total_inv = sum(inv_vols.values()) or 1
    raw     = {sym: v / total_inv for sym, v in inv_vols.items()}
    clipped = {sym: max(eq_w * VOL_MIN_WEIGHT, min(eq_w * VOL_MAX_WEIGHT, w))
               for sym, w in raw.items()}
    total_clip = sum(clipped.values()) or 1
    return {sym: w / total_clip for sym, w in clipped.items()}, daily_stds


def add_weights_to_top(top_df, close_df, today, sleeve_cfg, regime_exposure):
    if top_df.empty: return top_df
    weights, daily_stds = calc_atr_weights(top_df, close_df, today)
    capital = sleeve_cfg.get("capital", 0)
    n       = len(top_df)
    eq_wt   = 1.0 / n if n > 0 else 0
    rows    = []
    for _, row in top_df.iterrows():
        sym = row["Yahoo"]
        wt  = weights.get(sym, eq_wt)
        std = daily_stds.get(sym, 0.0)
        dep = capital * regime_exposure * wt if capital > 0 else None
        px  = float(row.get("Price", 0) or 0)
        qty = int(dep / px) if (dep and px > 0) else None
        row = row.copy()
        row["Daily_Std%"] = round(std, 3)
        row["Equal_Wt%"]  = round(eq_wt * 100, 2)
        row["ATR_Wt%"]    = round(wt * 100, 2)
        row["Deploy_Cap"] = round(dep, 0) if dep is not None else None
        row["Est_Qty"]    = qty
        rows.append(row)
    return pd.DataFrame(rows).reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
#  ❾  RS SCORER  — uses global ANALYSIS_DATE so past-date analysis works
# ══════════════════════════════════════════════════════════════════════════════

def score_today(sleeve_cfg, universe, cache):
    global ANALYSIS_DATE
    market  = sleeve_cfg["market"]
    rs_w    = sleeve_cfg["rs_weights"]
    top_n   = sleeve_cfg["top_n"]
    idx_sym = INDIA_INDEX if market == "INDIA" else US_INDEX
    sectors = INDIA_SECTORS if market == "INDIA" else US_SECTORS

    # ── Date range based on ANALYSIS_DATE (supports backtesting) ──────────
    end_dt   = ANALYSIS_DATE.strftime("%Y-%m-%d")
    start_dt = (ANALYSIS_DATE - timedelta(days=LOOKBACK_DAYS)).strftime("%Y-%m-%d")

    ref_syms = [idx_sym] + list(sectors.values())
    all_syms = list(dict.fromkeys(ref_syms + universe["Yahoo"].tolist()))

    print(f"  Loading prices up to {end_dt} …")
    closes, lows, highs, vols = cache.get_all(all_syms, start_dt, end_dt)

    # ── KEY: filter ALL data to ANALYSIS_DATE (critical for back-testing) ──
    analysis_ts = pd.Timestamp(ANALYSIS_DATE)
    if not closes.empty:
        closes = closes[closes.index <= analysis_ts]
    if not lows.empty:
        lows   = lows[lows.index <= analysis_ts]
    if not highs.empty:
        highs  = highs[highs.index <= analysis_ts]
    if not vols.empty:
        vols   = vols[vols.index <= analysis_ts]

    if idx_sym not in closes.columns:
        print(f"  ERROR: Index {idx_sym} not found in cache!")
        return pd.DataFrame(), {}, pd.DataFrame(), pd.DataFrame()

    idx_s                    = closes[idx_sym].dropna()
    regime, exp_frac, r_vals = get_regime(idx_s)
    avail                    = [s for s in universe["Yahoo"] if s in closes.columns]
    close_df                 = closes[avail]
    vol_df = vols.reindex(columns=avail) if not vols.empty and avail else pd.DataFrame()
    today  = close_df.index[-1]

    print(f"  Scoring {len(avail)} stocks (cap tier: {sleeve_cfg['cap_tier']}) …")

    stk_ret  = close_df / close_df.shift(RS_DECISION_PERIOD) - 1
    idx_ret  = (idx_s / idx_s.shift(RS_DECISION_PERIOD) - 1).reindex(close_df.index, method="ffill")

    sec_ret, ind_ret = {}, {}
    for sec in universe["Sector"].unique():
        cols = [c for c in universe[universe["Sector"] == sec]["Yahoo"] if c in stk_ret.columns]
        if cols: sec_ret[sec] = stk_ret[cols].mean(axis=1)
    for ind in universe["Industry"].unique():
        cols = [c for c in universe[universe["Industry"] == ind]["Yahoo"] if c in stk_ret.columns]
        if cols: ind_ret[ind] = stk_ret[cols].mean(axis=1)

    rs_mat = {}
    for p in rs_w:
        idr       = idx_s / idx_s.shift(p)
        rs_mat[p] = (close_df / close_df.shift(p)).div(
                        idr.reindex(close_df.index, method="ffill"), axis=0) - 1

    rs_sum = pd.DataFrame(0.0, index=close_df.index, columns=close_df.columns)
    wt_sum = pd.DataFrame(0.0, index=close_df.index, columns=close_df.columns)
    for p, w in rs_w.items():
        valid   = rs_mat[p].notna()
        rs_sum += rs_mat[p].fillna(0) * w * valid
        wt_sum += w * valid
    rs_comp = rs_sum.div(wt_sum.replace(0, np.nan))

    divisor       = 1e7 if market == "INDIA" else 1e6
    turnover_roll = pd.DataFrame()
    if not vol_df.empty:
        aligned_vol   = vol_df.reindex(close_df.index).fillna(0)
        turnover_roll = (close_df * aligned_vol).rolling(VOLUME_LOOKBACK, min_periods=5).mean() / divisor

    if today not in stk_ret.index:
        today = stk_ret.index[-1]
    sr     = stk_ret.loc[today]
    ir     = float(idx_ret.loc[today]) if today in idx_ret.index else np.nan
    min_t  = MIN_TURNOVER_CR if market == "INDIA" else MIN_TURNOVER_USD

    rows = []
    for _, row in universe.iterrows():
        sym, sec, ind = row["Yahoo"], row["Sector"], row["Industry"]
        if sym not in close_df.columns: continue
        s_ret = sr.get(sym, np.nan)
        i_ret = ind_ret.get(ind, pd.Series(dtype=float))
        c_ret = sec_ret.get(sec, pd.Series(dtype=float))
        i_val = float(i_ret.loc[today]) if today in i_ret.index else np.nan
        c_val = float(c_ret.loc[today]) if today in c_ret.index else np.nan
        if pd.isna(s_ret) or pd.isna(i_val) or pd.isna(c_val) or pd.isna(ir): continue
        if STRICT_PEER_FILTER:
            if not (s_ret >= i_val and s_ret >= c_val and c_val > ir and i_val > ir): continue
        rs_score = float(rs_comp.at[today, sym]) if (
            today in rs_comp.index and sym in rs_comp.columns) else np.nan
        if pd.isna(rs_score) or rs_score < MIN_RS_SCORE: continue
        avg_turn = np.nan
        if not turnover_roll.empty and sym in turnover_roll.columns and today in turnover_roll.index:
            avg_turn = turnover_roll.at[today, sym]
            if not pd.isna(avg_turn) and avg_turn < min_t: continue
        try:
            cur_px = round(float(close_df.at[today, sym]), 2)
        except:
            continue
        rs_vals = {p: round(float(rs_mat[p].at[today, sym]) * 100, 2)
                   if (today in rs_mat[p].index and sym in rs_mat[p].columns) else np.nan
                   for p in rs_w}
        rows.append({
            "Symbol":    row["Symbol"],  "Yahoo":  sym,
            "Name":      row["Company Name"], "Sector": sec, "Industry": ind,
            "Price":     cur_px,
            "RS_Score":  round(rs_score * 100, 4),
            "RS_22d":    rs_vals.get(22,  np.nan),
            "RS_55d":    rs_vals.get(55,  np.nan),
            "RS_120d":   rs_vals.get(120, np.nan),
            "RS_252d":   rs_vals.get(252, np.nan),
            f"Sec_{RS_DECISION_PERIOD}d":   round(c_val * 100, 2),
            f"Ind_{RS_DECISION_PERIOD}d":   round(i_val * 100, 2),
            f"Idx_{RS_DECISION_PERIOD}d":   round(ir * 100, 2),
            "Avg_Turnover": round(avg_turn, 2) if not pd.isna(avg_turn) else np.nan,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df, {"regime": regime, "exposure": exp_frac, "vals": r_vals}, pd.DataFrame(), close_df

    df = df.sort_values("RS_Score", ascending=False).reset_index(drop=True)
    df.insert(0, "RS_Rank", df.index + 1)

    sec_cap  = max(1, int(top_n * SECTOR_CAP_PCT))
    sec_cnt  = {}
    top_rows = []
    for _, r in df.iterrows():
        sc = r["Sector"]
        if sec_cnt.get(sc, 0) < sec_cap:
            top_rows.append(r)
            sec_cnt[sc] = sec_cnt.get(sc, 0) + 1
        if len(top_rows) >= top_n:
            break
    top_df = pd.DataFrame(top_rows).reset_index(drop=True) if top_rows else pd.DataFrame()

    return df, {"regime": regime, "exposure": exp_frac, "vals": r_vals}, top_df, close_df


# ══════════════════════════════════════════════════════════════════════════════
#  ❿  HOLDINGS + ACTIONS  — uses ANALYSIS_DATE for Days_Held
# ══════════════════════════════════════════════════════════════════════════════

def load_holdings(sleeve_name):
    if not os.path.exists(HOLDINGS_CSV):
        return pd.DataFrame(columns=["Symbol", "Yahoo", "EntryDate", "EntryPrice", "Qty", "Sleeve"])
    df = pd.read_csv(HOLDINGS_CSV)
    df.columns = df.columns.str.strip()
    if "Sleeve" in df.columns:
        df = df[df["Sleeve"].astype(str).str.upper() == sleeve_name.upper()]
    return df


def compute_actions(holdings, top_df, all_scores_df, sleeve_cfg, regime_info):
    global ANALYSIS_DATE
    sl      = sleeve_cfg.get("stop_loss", 0)
    tp      = sleeve_cfg.get("take_profit", 0)
    top_set = set(top_df["Yahoo"].tolist()) if not top_df.empty else set()
    held_set = set(holdings["Yahoo"].tolist()) if not holdings.empty else set()

    wt_lookup = {}
    if not top_df.empty and "ATR_Wt%" in top_df.columns:
        for _, r in top_df.iterrows():
            wt_lookup[r["Yahoo"]] = {k: r.get(k) for k in
                ["ATR_Wt%", "Equal_Wt%", "Deploy_Cap", "Est_Qty",
                 "RS_Rank", "RS_Score", "Daily_Std%", "Sector"]}

    actions = []

    # SELLs
    for _, pos in holdings.iterrows():
        y       = pos["Yahoo"]
        reason  = "RANK_EXIT"
        entry   = float(pos.get("EntryPrice", 0) or 0)
        cur_row = all_scores_df[all_scores_df["Yahoo"] == y] if not all_scores_df.empty else pd.DataFrame()
        cur_px  = float(cur_row.iloc[0]["Price"]) if len(cur_row) > 0 else None
        if cur_px and entry > 0:
            ret = (cur_px - entry) / entry
            if sl > 0 and ret <= -sl:  reason = f"STOP_LOSS ({ret*100:.1f}%)"
            elif tp > 0 and ret >= tp: reason = f"TARGET_HIT ({ret*100:.1f}%)"
        if y not in top_set:
            entry_date = pd.to_datetime(pos.get("EntryDate", ANALYSIS_DATE)).date()
            actions.append({
                "Action": "SELL", "Symbol": pos.get("Symbol", y), "Yahoo": y,
                "Reason": reason, "Entry_Price": entry, "Current_Price": cur_px,
                "Return_%": round((cur_px / entry - 1) * 100, 1) if cur_px and entry else None,
                "Days_Held": (ANALYSIS_DATE - entry_date).days,
            })

    # BUYs
    if not top_df.empty:
        for _, r in top_df.iterrows():
            if r["Yahoo"] not in held_set:
                wt = wt_lookup.get(r["Yahoo"], {})
                actions.append({
                    "Action": "BUY", "Symbol": r["Symbol"], "Yahoo": r["Yahoo"],
                    "Reason": "RS_ENTRY", "RS_Rank": r.get("RS_Rank"),
                    "RS_Score": r.get("RS_Score"), "Sector": r.get("Sector"),
                    "Current_Price": r.get("Price"), **wt,
                })

    # HOLDs
    for _, pos in holdings.iterrows():
        y = pos["Yahoo"]
        if y in top_set:
            wt      = wt_lookup.get(y, {})
            cur_row = all_scores_df[all_scores_df["Yahoo"] == y] if not all_scores_df.empty else pd.DataFrame()
            cur_px  = float(cur_row.iloc[0]["Price"]) if len(cur_row) > 0 else None
            entry   = float(pos.get("EntryPrice", 0) or 0)
            actions.append({
                "Action": "HOLD", "Symbol": pos.get("Symbol", y), "Yahoo": y,
                "Reason": "STILL_IN_TOP",
                "RS_Rank_New": wt.get("RS_Rank"), "RS_Score": wt.get("RS_Score"),
                "Sector": wt.get("Sector"), "Entry_Price": entry,
                "Current_Price": cur_px,
                "Return_%": round((cur_px / entry - 1) * 100, 1) if cur_px and entry else None,
                **{k: wt.get(k) for k in ["ATR_Wt%", "Equal_Wt%", "Deploy_Cap", "Est_Qty", "Daily_Std%"]},
            })

    return pd.DataFrame(actions)


# ══════════════════════════════════════════════════════════════════════════════
#  ⓫  SNAPSHOT SAVE / LOAD / COMPARE  (NEW in v2.0)
# ══════════════════════════════════════════════════════════════════════════════

def save_snapshot(sleeve_name, top_df, regime_info, analysis_date):
    """Save current top picks as JSON for future month-over-month comparison."""
    Path(SNAPSHOT_DIR).mkdir(parents=True, exist_ok=True)
    snap = {
        "sleeve":   sleeve_name,
        "date":     str(analysis_date),
        "regime":   regime_info.get("regime", "?"),
        "exposure": regime_info.get("exposure", 1.0),
        "stocks":   [],
    }
    if not top_df.empty:
        for _, r in top_df.iterrows():
            def _safe(v, cast=None):
                if v is None: return None
                try:
                    return cast(v) if (cast and not pd.isna(v)) else None
                except:
                    return None
            snap["stocks"].append({
                "Symbol":   str(r.get("Symbol", "")),
                "Yahoo":    str(r.get("Yahoo", "")),
                "Sector":   str(r.get("Sector", "")),
                "RS_Rank":  _safe(r.get("RS_Rank"), int),
                "RS_Score": _safe(r.get("RS_Score"), float),
                "Price":    _safe(r.get("Price"), float),
                "ATR_Wt%":  _safe(r.get("ATR_Wt%"), float),
                "Equal_Wt%":_safe(r.get("Equal_Wt%"), float),
                "Est_Qty":  _safe(r.get("Est_Qty"), int),
            })
    date_str = str(analysis_date).replace("-", "")
    fname = os.path.join(SNAPSHOT_DIR, f"RS_Snap_{sleeve_name}_{date_str}.json")
    with open(fname, "w") as f:
        json.dump(snap, f, indent=2)
    print(f"  Snapshot saved → {fname}")
    return fname


def load_prev_snapshot(sleeve_name, analysis_date, max_lookback_days=90):
    """Find and load the most recent snapshot BEFORE analysis_date (within max_lookback_days)."""
    snap_dir = Path(SNAPSHOT_DIR)
    if not snap_dir.exists():
        return None
    pattern = str(snap_dir / f"RS_Snap_{sleeve_name}_*.json")
    files   = sorted(glob.glob(pattern))

    best = None
    for fpath in files:
        fname = os.path.basename(fpath)
        # filename: RS_Snap_{sleeve}_{YYYYMMDD}.json
        # sleeve name may contain underscores, so extract date from end
        try:
            date_part = fname.replace(".json", "").split("_")[-1]
            snap_date = date(int(date_part[:4]), int(date_part[4:6]), int(date_part[6:8]))
        except:
            continue
        days_back = (analysis_date - snap_date).days
        if 0 < days_back <= max_lookback_days:
            best = (fpath, snap_date)

    if best is None:
        return None
    with open(best[0]) as f:
        data = json.load(f)
    print(f"  Previous snapshot found: {best[0]} ({best[1]})")
    return data


def compare_snapshots(current_top_df, prev_snap, sleeve_cfg):
    """
    Diff current top picks against previous snapshot.
    Returns a DataFrame with Status, Action, rank changes, weight changes, qty adjustments.
    """
    if prev_snap is None or current_top_df.empty:
        return pd.DataFrame()

    prev_stocks = {s["Symbol"]: s for s in prev_snap.get("stocks", [])}
    curr_stocks = {}
    for _, r in current_top_df.iterrows():
        curr_stocks[str(r["Symbol"])] = r

    prev_syms = set(prev_stocks.keys())
    curr_syms = set(curr_stocks.keys())
    all_syms  = prev_syms | curr_syms

    rows = []
    for sym in sorted(all_syms):
        in_prev = sym in prev_syms
        in_curr = sym in curr_syms
        p = prev_stocks.get(sym, {})
        c = curr_stocks.get(sym, {})

        if in_prev and in_curr:
            status = "HOLD"
        elif in_curr:
            status = "NEW"
        else:
            status = "REMOVED"

        def _flt(v):
            try:
                return float(v) if (v is not None and not pd.isna(v)) else None
            except:
                return None
        def _int(v):
            try:
                return int(v) if (v is not None and not pd.isna(v)) else None
            except:
                return None

        prev_rank  = _int(p.get("RS_Rank"))
        curr_rank  = _int(c.get("RS_Rank")) if in_curr else None
        rank_chg   = (curr_rank - prev_rank) if (prev_rank and curr_rank) else None

        prev_score = _flt(p.get("RS_Score"))
        curr_score = _flt(c.get("RS_Score")) if in_curr else None

        prev_wt    = _flt(p.get("ATR_Wt%"))
        curr_wt    = _flt(c.get("ATR_Wt%")) if in_curr else None
        wt_chg     = round(curr_wt - prev_wt, 2) if (prev_wt and curr_wt) else None

        prev_qty   = _int(p.get("Est_Qty"))
        curr_qty   = _int(c.get("Est_Qty")) if in_curr else None
        qty_chg    = (curr_qty - prev_qty) if (curr_qty and prev_qty) else None

        # Action recommendation
        if status == "NEW":
            action = "BUY"
        elif status == "REMOVED":
            action = "SELL"
        elif wt_chg is not None and abs(wt_chg) >= 1.0:
            direction = "ADD" if wt_chg > 0 else "REDUCE"
            action = f"{direction} ({wt_chg:+.1f}%)"
        else:
            action = "NO CHANGE"

        rows.append({
            "Status":     status,
            "Action":     action,
            "Symbol":     sym,
            "Sector":     str(c.get("Sector", p.get("Sector", "?"))) if isinstance(c, (pd.Series, dict)) else "?",
            "Prev_Rank":  prev_rank,
            "Curr_Rank":  curr_rank,
            "Rank_Δ":     rank_chg,
            "Prev_RS":    round(prev_score, 2) if prev_score else None,
            "Curr_RS":    round(curr_score, 2) if curr_score else None,
            "Prev_Wt%":   prev_wt,
            "Curr_Wt%":   curr_wt,
            "Wt_Δ%":      wt_chg,
            "Prev_Qty":   prev_qty,
            "Curr_Qty":   curr_qty,
            "Qty_Δ":      qty_chg,
            "Prev_Date":  prev_snap.get("date", "?"),
        })

    if not rows:
        return pd.DataFrame()

    order = {"NEW": 0, "HOLD": 1, "REMOVED": 2}
    df = pd.DataFrame(rows)
    df["_ord"] = df["Status"].map(order)
    df = df.sort_values(["_ord", "Curr_Rank"]).drop("_ord", axis=1).reset_index(drop=True)
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  ⓬  CONSOLE PRINT
# ══════════════════════════════════════════════════════════════════════════════

def _fmt_cap(val, market):
    if val is None: return "—"
    sym = "₹" if market == "INDIA" else "$"
    if val >= 1_00_000: return f"{sym}{val/1_00_000:.1f}L"
    if val >= 1_000:    return f"{sym}{val/1_000:.1f}K"
    return f"{sym}{int(val)}"


def print_report(sleeve_name, cfg, top_df, actions_df, regime_info):
    global ANALYSIS_DATE
    SEP = "═" * 78
    mkt = cfg.get("market", "INDIA")
    cap = cfg.get("capital", 0)
    cur = "₹" if mkt == "INDIA" else "$"

    date_label = str(ANALYSIS_DATE)
    if ANALYSIS_DATE != date.today():
        date_label += f"  ⏪ BACKTEST"

    print(f"\n{SEP}")
    print(f"  Sleeve {sleeve_name} — {cfg['name']}  |  {cfg['cap_tier']}")
    print(f"  {cfg['description']}")
    print(f"  Analysis Date: {date_label}  |  TopN: {cfg['top_n']}  |  Freq: {cfg['rebalance']}")
    print(f"  RS Weights: 22d={cfg['rs_weights'].get(22,0)*100:.0f}%  "
          f"55d={cfg['rs_weights'].get(55,0)*100:.0f}%  "
          f"120d={cfg['rs_weights'].get(120,0)*100:.0f}%  "
          f"252d={cfg['rs_weights'].get(252,0)*100:.0f}%")
    print(SEP)

    regime  = regime_info.get("regime", "?")
    exp_pct = regime_info.get("exposure", 1.0)
    rvals   = regime_info.get("vals", {})
    icons   = {"BULL": "[BULL]", "CAUTION": "[CAUTION]", "BEAR": "[BEAR]"}
    dep_cap = cap * exp_pct if cap else None

    print(f"\n  {icons.get(regime, '?')} REGIME: {regime}  |  Deploy {int(exp_pct*100)}% "
          + (f"= {_fmt_cap(dep_cap, mkt)} of {_fmt_cap(cap, mkt)}" if dep_cap else ""))
    if rvals:
        print(f"     Index: {rvals.get('Index','?')}  |  "
              f"EMA100: {rvals.get('EMA100','?')}  |  EMA200: {rvals.get('EMA200','?')}")
    print(f"     Sizing: {'ATR inverse-vol' if USE_VOL_SIZING else 'Equal weight'}  "
          f"|  Floor: {VOL_MIN_WEIGHT}x eq  |  Cap: {VOL_MAX_WEIGHT}x eq")

    if not actions_df.empty:
        sells = actions_df[actions_df["Action"] == "SELL"]
        buys  = actions_df[actions_df["Action"] == "BUY"]
        holds = actions_df[actions_df["Action"] == "HOLD"]

        if not sells.empty:
            print(f"\n  SELL ({len(sells)} positions):")
            print(f"  {'Symbol':<12} {'Reason':<28} {'Entry':>8} {'Now':>8} {'P&L':>8}  Days")
            print("  " + "-" * 70)
            for _, r in sells.iterrows():
                ret_s = f"{r['Return_%']:+.1f}%" if r.get("Return_%") is not None else "N/A"
                print(f"  {str(r.get('Symbol','?')):<12} {str(r.get('Reason','?')):<28} "
                      f"{str(r.get('Entry_Price','?')):>8} {str(r.get('Current_Price','?')):>8} "
                      f"{ret_s:>8}  {str(r.get('Days_Held','?'))}")

        if not buys.empty:
            has_cap = ("Deploy_Cap" in buys.columns and buys["Deploy_Cap"].notna().any())
            print(f"\n  BUY ({len(buys)} new positions):")
            if has_cap:
                print(f"  {'Rk':<4} {'Symbol':<12} {'Sector':<16} {'RS':>7} "
                      f"{'Price':>8} {'Std%':>6} {'EqWt':>6} {'ATRWt':>6} {'Capital':>9} {'Qty':>6}")
                print("  " + "-" * 78)
                for _, r in buys.iterrows():
                    print(f"  {str(r.get('RS_Rank','?')):<4} {str(r.get('Symbol','?')):<12} "
                          f"{str(r.get('Sector','?'))[:15]:<16} "
                          f"{r.get('RS_Score',0):>7.2f} {r.get('Current_Price',0):>8.2f} "
                          f"{r.get('Daily_Std%',0):>5.2f}% "
                          f"{r.get('Equal_Wt%',0):>5.1f}% {r.get('ATR_Wt%',0):>5.1f}% "
                          f"{_fmt_cap(r.get('Deploy_Cap'), mkt):>9} "
                          f"{str(int(r['Est_Qty'])) if r.get('Est_Qty') else '—':>6}")
            else:
                print(f"  {'Rk':<4} {'Symbol':<12} {'Sector':<16} {'RS':>7} "
                      f"{'Price':>8} {'Std%':>6} {'EqWt':>6} {'ATRWt':>6}")
                print("  " + "-" * 66)
                for _, r in buys.iterrows():
                    print(f"  {str(r.get('RS_Rank','?')):<4} {str(r.get('Symbol','?')):<12} "
                          f"{str(r.get('Sector','?'))[:15]:<16} "
                          f"{r.get('RS_Score',0):>7.2f} {r.get('Current_Price',0):>8.2f} "
                          f"{r.get('Daily_Std%',0):>5.2f}% "
                          f"{r.get('Equal_Wt%',0):>5.1f}% {r.get('ATR_Wt%',0):>5.1f}%")

        if not holds.empty:
            has_cap = ("Deploy_Cap" in holds.columns and holds["Deploy_Cap"].notna().any())
            print(f"\n  HOLD ({len(holds)} positions — no action):")
            fmt = (f"  {'Symbol':<12} {'NewRk':>6} {'P&L':>8} {'Std%':>6} "
                   f"{'EqWt':>6} {'ATRWt':>6} {'Capital':>9}")
            print(fmt if has_cap else fmt.replace(" {'Capital':>9}", ""))
            print("  " + "-" * (70 if has_cap else 58))
            for _, r in holds.iterrows():
                ret_s = f"{r['Return_%']:+.1f}%" if r.get("Return_%") is not None else "—"
                line  = (f"  {str(r.get('Symbol','?')):<12} "
                         f"#{str(r.get('RS_Rank_New','?')):<5} {ret_s:>8} "
                         f"{r.get('Daily_Std%',0):>5.2f}% "
                         f"{r.get('Equal_Wt%',0):>5.1f}% {r.get('ATR_Wt%',0):>5.1f}%")
                if has_cap: line += f" {_fmt_cap(r.get('Deploy_Cap'), mkt):>9}"
                print(line)

    if not top_df.empty:
        has_cap = ("Deploy_Cap" in top_df.columns and top_df["Deploy_Cap"].notna().any())
        total_d = top_df["Deploy_Cap"].sum() if has_cap else None
        print(f"\n  FULL PORTFOLIO — {cfg['cap_tier']}:")
        if total_d:
            print(f"  Total deployed: {_fmt_cap(total_d, mkt)}  "
                  f"(regime {int(exp_pct*100)}% of {_fmt_cap(cap, mkt)})")
        hdr = (f"  {'Rk':<4} {'Symbol':<12} {'Sector':<16} {'RS':>7} "
               f"{'Price':>8} {'Std%':>6} {'EqWt':>6} {'ATRWt':>6}")
        if has_cap: hdr += f" {'Capital':>9} {'Qty':>6}"
        print(hdr)
        print("  " + "─" * (78 if has_cap else 60))
        for _, r in top_df.iterrows():
            line = (f"  {int(r['RS_Rank']):<4} {str(r['Symbol']):<12} "
                    f"{str(r.get('Sector','?'))[:15]:<16} "
                    f"{r['RS_Score']:>7.2f} {r['Price']:>8.2f} "
                    f"{r.get('Daily_Std%',0):>5.2f}% "
                    f"{r.get('Equal_Wt%',0):>5.1f}% {r.get('ATR_Wt%',0):>5.1f}%")
            if has_cap:
                line += (f" {_fmt_cap(r.get('Deploy_Cap'), mkt):>9} "
                         f"{str(int(r['Est_Qty'])) if r.get('Est_Qty') else '—':>6}")
            print(line)
        print(f"  Weight total: {top_df['ATR_Wt%'].sum():.1f}%  (should be 100.0%)")
    print(f"\n{SEP}\n")


# ══════════════════════════════════════════════════════════════════════════════
#  ⓭  EXCEL OUTPUT  — adds Monthly Comparison sheet (NEW in v2.0)
# ══════════════════════════════════════════════════════════════════════════════

def _F(h):   return PatternFill("solid", fgColor=h)
def _FN(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Arial", italic=italic)
def _AL(h="center"): return Alignment(horizontal=h, vertical="center")
_THIN = Side(style="thin", color="E0E0E0")
_BD   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hdr(cell, text, bg="1F2D54"):
    cell.value     = text
    cell.fill      = _F(bg)
    cell.font      = _FN(bold=True, color="FFFFFF", size=9)
    cell.alignment = _AL("center")
    cell.border    = _BD


def _xl_cap(val):
    if val is None: return None
    try:   return int(round(val, 0))
    except: return None


def _auto_w(ws, max_w=22):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=6)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 2, max_w)


def write_sleeve_sheets(wb, sleeve_name, cfg, all_scores_df, top_df, actions_df,
                        regime_info, comparison_df=None):
    global ANALYSIS_DATE
    mkt     = cfg.get("market", "INDIA")
    cap     = cfg.get("capital", 0)
    cur     = "₹" if mkt == "INDIA" else "$"
    today_s = ANALYSIS_DATE.strftime("%Y-%m-%d")
    regime  = regime_info.get("regime", "?")
    exp_pct = regime_info.get("exposure", 1.0)
    rvals   = regime_info.get("vals", {})
    tier_hdr_color = {
        "A": "1F2D54", "B": "0F6E56", "C": "854F0B", "D": "5F5E5A",
        "US_A": "185FA5", "US_B": "0F6E8A", "US_C": "6B2E85", "US_ETF": "7A5200",
    }.get(sleeve_name, "1F2D54")

    backtest_label = f" (BACKTEST)" if ANALYSIS_DATE != date.today() else ""

    # Sheet name prefix so all sleeve sheets are clearly grouped
    pfx = f"{sleeve_name}_"

    # ── Sleeve D (Liquid) — single info sheet, no scoring data ──────────────
    if not cfg.get("universe"):
        ws = wb.create_sheet(f"{pfx}Info")
        ws.sheet_view.showGridLines = False
        ws["A1"] = f"Sleeve {sleeve_name}: {cfg['name']} — {cfg['cap_tier']}"
        ws["A1"].font = _FN(bold=True, size=14, color=tier_hdr_color)
        ws["A2"] = f"Analysis Date: {today_s}{backtest_label}"
        ws["A2"].font = _FN(italic=True, size=10, color="666666")
        info_rows = [
            ("", ""),
            ("Purpose",        "Liquid buffer / bear-market dry powder"),
            ("Instrument",     "Liquid Bees ETF  (or short-term FD)"),
            ("Expected Yield", "~6-7% p.a."),
            ("Capital",        f"{cur}{cap:,.0f}"),
            ("", ""),
            ("BULL regime",    "Deploy capital back into Sleeves A, B, C as directed"),
            ("CAUTION regime", "Hold 50% here; reduce A/B/C exposure"),
            ("BEAR regime",    "Receive capital released from A, B, C; park here"),
            ("", ""),
            ("RS Scoring",     "Not applicable — no momentum score needed for cash/liquid"),
            ("Rebalance",      cfg.get("rebalance", "as_needed").replace("_", " ").title()),
        ]
        for i, (label, value) in enumerate(info_rows, 4):
            cl = ws.cell(i, 1, label)
            cv = ws.cell(i, 2, value)
            cl.font = _FN(bold=True, size=10, color=tier_hdr_color)
            cv.font = _FN(size=10)
            cl.alignment = _AL("left")
            cv.alignment = _AL("left")
            if label:
                cl.border = _BD
                cv.border = _BD
                if i % 2 == 0:
                    cl.fill = _F("F5F5F5")
                    cv.fill = _F("F5F5F5")
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 55
        return wb

    # ── Sheet 1: Action List ──────────────────────────────────────────────────
    ws1 = wb.create_sheet(f"{pfx}Actions")
    ws1.sheet_view.showGridLines = False
    ws1["A1"] = f"RS TRACKER v2.0 — Sleeve {sleeve_name}: {cfg['name']} — {today_s}{backtest_label}"
    ws1["A1"].font = _FN(bold=True, size=12, color=tier_hdr_color)
    ws1["A2"] = f"Cap Tier: {cfg['cap_tier']}  |  {cfg['description']}"
    ws1["A2"].font = _FN(italic=True, size=9, color="666666")
    ws1["A3"] = (f"RS Weights: 22d={cfg['rs_weights'].get(22,0)*100:.0f}%  "
                 f"55d={cfg['rs_weights'].get(55,0)*100:.0f}%  "
                 f"120d={cfg['rs_weights'].get(120,0)*100:.0f}%  "
                 f"252d={cfg['rs_weights'].get(252,0)*100:.0f}%  "
                 f"|  Sizing: {'ATR' if USE_VOL_SIZING else 'Equal'}")
    ws1["A3"].font = _FN(size=9, color="444444")

    rc = {"BULL": "1A6B3C", "CAUTION": "854F0B", "BEAR": "7F1D1D"}
    ws1["A5"] = f"REGIME: {regime}  —  Deploy {int(exp_pct*100)}% of sleeve capital"
    ws1["A5"].font = _FN(bold=True, size=11, color=rc.get(regime, "000000"))
    if cap:
        ws1["D5"] = f"Sleeve: {cur}{cap:,.0f}   Deployed: {cur}{cap*exp_pct:,.0f}"
        ws1["D5"].font = _FN(bold=True, size=10)
    if rvals:
        ws1["G5"] = (f"Index: {rvals.get('Index','?')}  "
                     f"EMA100: {rvals.get('EMA100','?')}  EMA200: {rvals.get('EMA200','?')}")
        ws1["G5"].font = _FN(size=9, color="555555")

    if not actions_df.empty:
        for action_type, lbl_bg in [("SELL", "B91C1C"), ("BUY", "1A6B3C"), ("HOLD", "2E4482")]:
            sub = actions_df[actions_df["Action"] == action_type]
            if sub.empty: continue
            row = ws1.max_row + 2
            ws1.cell(row, 1, f"  {action_type} — {len(sub)} positions  ")
            ws1.cell(row, 1).font = _FN(bold=True, size=10, color="FFFFFF")
            ws1.cell(row, 1).fill = _F(lbl_bg)
            ws1.merge_cells(f"A{row}:M{row}")
            row += 1

            if action_type == "SELL":
                cols = ["Action", "Symbol", "Reason", "Entry_Price", "Current_Price",
                        "Return_%", "Days_Held"]
            elif action_type == "BUY":
                cols = ["Action", "Symbol", "Sector", "RS_Rank", "RS_Score",
                        "Current_Price", "Daily_Std%", "Equal_Wt%", "ATR_Wt%",
                        "Demo_Capital", "Demo_Est_Qty"]
            else:
                cols = ["Action", "Symbol", "Sector", "RS_Rank_New", "RS_Score",
                        "Entry_Price", "Current_Price", "Return_%", "Daily_Std%",
                        "Equal_Wt%", "ATR_Wt%", "Demo_Capital", "Demo_Est_Qty"]
            cols = [c for c in cols if c in sub.columns]

            for j, c in enumerate(cols, 1):
                lbl = (c.replace("Demo_Capital", f"Capital ({cur}{SLEEVE_DEMO_CAPITAL:,}/sleeve)")
                         .replace("Demo_Est_Qty", "Est Qty")
                         .replace("Deploy_Cap", f"Capital {cur}")
                         .replace("_", " "))
                _hdr(ws1.cell(row, j), lbl, bg="444444")
            row += 1

            for _, r in sub.iterrows():
                # Compute demo capital & qty from ATR_Wt% × SLEEVE_DEMO_CAPITAL
                atr_wt  = r.get("ATR_Wt%")
                cur_px  = r.get("Current_Price")
                try:
                    demo_cap = round(float(atr_wt) / 100.0 * SLEEVE_DEMO_CAPITAL, 2) if atr_wt is not None else None
                except Exception:
                    demo_cap = None
                try:
                    demo_qty = int(demo_cap / float(cur_px)) if (demo_cap and cur_px and float(cur_px) > 0) else None
                except Exception:
                    demo_qty = None

                for j, c in enumerate(cols, 1):
                    if c == "Demo_Capital":
                        val = demo_cap
                    elif c == "Demo_Est_Qty":
                        val = demo_qty
                    else:
                        val = r.get(c)
                    if isinstance(val, float) and np.isnan(val): val = None
                    if c == "Deploy_Cap": val = _xl_cap(val)
                    cell = ws1.cell(row, j, val)
                    cell.alignment = _AL("center")
                    cell.font      = _FN(size=9)
                    cell.border    = _BD
                    if c == "Action":
                        cell.fill = _F({"SELL": "FECDD3", "BUY": "D1FAE5", "HOLD": "DBEAFE"}.get(str(val), "FFFFFF"))
                    elif c == "Return_%":
                        if isinstance(val, (int, float)) and val is not None:
                            cell.fill = _F("D1FAE5" if val >= 0 else "FECDD3")
                    elif c in ("ATR_Wt%", "Demo_Capital", "Deploy_Cap"):
                        cell.font = _FN(size=9, bold=True, color=tier_hdr_color)
                    elif c == "Demo_Est_Qty":
                        cell.font = _FN(size=9, bold=True, color="1A6B3C")
                    elif c == "Reason" and "STOP" in str(val):
                        cell.fill = _F("FECDD3")
                    elif c == "Reason" and "TARGET" in str(val):
                        cell.fill = _F("D1FAE5")
                row += 1
    _auto_w(ws1)

    # ── Sheet 2: Full RS Rankings ─────────────────────────────────────────────
    ws2 = wb.create_sheet(f"{pfx}RS Rankings")
    ws2.sheet_view.showGridLines = False
    ws2["A1"] = (f"RS RANKINGS — {cfg['cap_tier']} — {len(all_scores_df)} passed filter — {today_s}")
    ws2["A1"].font = _FN(bold=True, size=11, color=tier_hdr_color)
    ws2["A2"] = f"Top {cfg['top_n']} (after sector cap) highlighted | RS scheme: {cfg['rs_weights']}"
    ws2["A2"].font = _FN(italic=True, size=9, color="666666")

    top_set  = set(top_df["Yahoo"].tolist()) if not top_df.empty else set()
    rank_cols = ["RS_Rank", "Symbol", "Name", "Sector", "Price", "RS_Score",
                 "RS_22d", "RS_55d", "RS_120d", "RS_252d",
                 f"Sec_{RS_DECISION_PERIOD}d", f"Ind_{RS_DECISION_PERIOD}d", "Avg_Turnover"]
    rank_cols = [c for c in rank_cols if c in all_scores_df.columns]
    yahoo_col = all_scores_df["Yahoo"] if "Yahoo" in all_scores_df.columns else pd.Series()
    START = 4
    for j, c in enumerate(rank_cols, 1):
        _hdr(ws2.cell(START, j), c, bg=tier_hdr_color)
    for i, (idx, row) in enumerate(all_scores_df[rank_cols].iterrows()):
        xl     = START + 1 + i
        in_top = yahoo_col.loc[idx] in top_set if idx in yahoo_col.index else False
        for j, val in enumerate(row, 1):
            v = None if isinstance(val, float) and np.isnan(val) else val
            c = ws2.cell(xl, j, v)
            c.alignment = _AL("center"); c.border = _BD
            c.font = _FN(size=9, bold=in_top)
            if in_top:    c.fill = _F("D1FAE5")
            elif i % 2 == 0: c.fill = _F("F8FAFF")
    _auto_w(ws2)
    ws2.freeze_panes = f"A{START+1}"

    # ── Sheet 3: Holdings Template ────────────────────────────────────────────
    ws3 = wb.create_sheet(f"{pfx}Holdings")
    ws3.sheet_view.showGridLines = False
    ws3["A1"] = "COPY THIS TO my_holdings.csv AFTER EXECUTING TRADES"
    ws3["A1"].font = _FN(bold=True, size=11, color="1A6B3C")
    ws3["A2"] = (f"Sleeve: {sleeve_name} ({cfg['cap_tier']}) | "
                 f"Regime: {regime} | Deploy: {int(exp_pct*100)}% | {today_s}{backtest_label}")
    ws3["A2"].font = _FN(italic=True, size=9, color="666666")

    templ_cols = ["Symbol", "Yahoo", "EntryDate", "EntryPrice", "Qty", "Sleeve",
                  "ATR_Wt%", "Equal_Wt%", "Daily_Std%", f"Capital_{cur}", "Est_Qty"]
    for j, c in enumerate(templ_cols, 1):
        _hdr(ws3.cell(4, j), c.replace("_", " "), bg="1A6B3C")
    if not top_df.empty:
        for i, (_, r) in enumerate(top_df.iterrows(), 5):
            vals = [r.get("Symbol", ""), r.get("Yahoo", ""), today_s, r.get("Price", ""),
                    r.get("Est_Qty", ""), sleeve_name, r.get("ATR_Wt%", ""),
                    r.get("Equal_Wt%", ""), r.get("Daily_Std%", ""),
                    _xl_cap(r.get("Deploy_Cap")), r.get("Est_Qty", "")]
            for j, v in enumerate(vals, 1):
                cell = ws3.cell(i, j, v)
                cell.font      = _FN(size=9, bold=(j == 7), color="1A6B3C" if j == 7 else "000000")
                cell.alignment = _AL("center")
                cell.border    = _BD
                if i % 2 == 0: cell.fill = _F("F0FFF4")
        sr = ws3.max_row + 1
        ws3.cell(sr, 1, "TOTAL").font = _FN(bold=True)
        ws3.cell(sr, 7, round(top_df["ATR_Wt%"].sum(), 1)).font = _FN(bold=True, color="1A6B3C")
        if "Deploy_Cap" in top_df.columns:
            ws3.cell(sr, 10, _xl_cap(top_df["Deploy_Cap"].sum())).font = _FN(bold=True, color=tier_hdr_color)
    _auto_w(ws3)

    # ── Sheet 4: Monthly Comparison  (NEW in v2.0) ───────────────────────────
    if comparison_df is not None and not comparison_df.empty:
        ws4 = wb.create_sheet(f"{pfx}Comparison")
        ws4.sheet_view.showGridLines = False

        prev_date = comparison_df["Prev_Date"].iloc[0] if "Prev_Date" in comparison_df.columns else "?"
        ws4["A1"] = f"MONTH-OVER-MONTH COMPARISON — Sleeve {sleeve_name} — {prev_date}  →  {today_s}"
        ws4["A1"].font = _FN(bold=True, size=12, color=tier_hdr_color)
        ws4["A2"] = ("GREEN = new entry (BUY) · RED = exited (SELL) · BLUE = held · "
                     "Bold italic = qty adjustment needed")
        ws4["A2"].font = _FN(italic=True, size=9, color="666666")

        # Summary counts
        n_new  = len(comparison_df[comparison_df["Status"] == "NEW"])
        n_sold = len(comparison_df[comparison_df["Status"] == "REMOVED"])
        n_held = len(comparison_df[comparison_df["Status"] == "HOLD"])
        n_adj  = len(comparison_df[comparison_df["Action"].str.contains("ADD|REDUCE", na=False)])
        ws4["A3"] = (f"Summary: {n_new} NEW  |  {n_sold} REMOVED  |  "
                     f"{n_held} HELD  |  {n_adj} need qty adjustment")
        ws4["A3"].font = _FN(bold=True, size=10, color="333333")

        comp_cols = ["Status", "Action", "Symbol", "Sector", "Prev_Rank", "Curr_Rank",
                     "Rank_Δ", "Prev_RS", "Curr_RS", "Prev_Wt%", "Curr_Wt%", "Wt_Δ%",
                     "Prev_Qty", "Curr_Qty", "Qty_Δ"]
        comp_cols = [c for c in comp_cols if c in comparison_df.columns]

        START4 = 5
        for j, c in enumerate(comp_cols, 1):
            _hdr(ws4.cell(START4, j), c, bg=tier_hdr_color)

        status_bg = {"NEW": "D1FAE5", "REMOVED": "FECDD3", "HOLD": "EFF6FF"}
        action_clr = {"BUY": "166534", "SELL": "991B1B"}

        for i, (_, row) in enumerate(comparison_df.iterrows()):
            xl  = START4 + 1 + i
            st  = row.get("Status", "HOLD")
            act = str(row.get("Action", ""))
            bg  = status_bg.get(st, "FFFFFF")
            needs_adj = "ADD" in act or "REDUCE" in act

            for j, c in enumerate(comp_cols, 1):
                val = row.get(c)
                if isinstance(val, float) and np.isnan(val): val = None
                cell = ws4.cell(xl, j, val)
                cell.alignment = _AL("center")
                cell.border    = _BD
                cell.fill      = _F(bg)
                cell.font      = _FN(size=9, bold=needs_adj, italic=needs_adj,
                                     color=action_clr.get(act, "000000") if j <= 2 else "000000")
                # Rank change colouring
                if c == "Rank_Δ" and isinstance(val, (int, float)) and val is not None:
                    cell.fill = _F("D1FAE5" if val < 0 else ("FECDD3" if val > 0 else bg))
                # Wt change colouring
                if c == "Wt_Δ%" and isinstance(val, (int, float)) and val is not None:
                    cell.fill = _F("D1FAE5" if val > 0 else ("FECDD3" if val < 0 else bg))
        _auto_w(ws4)
        ws4.freeze_panes = f"A{START4+1}"

    return wb


def save_combined_excel(wb, outpath):
    """Save the combined workbook that was built by write_sleeve_sheets calls."""
    wb.save(outpath)
    return outpath


# ══════════════════════════════════════════════════════════════════════════════
#  ⓮  MAIN RUNNER
# ══════════════════════════════════════════════════════════════════════════════

def run_sleeve(sleeve_name):
    global ANALYSIS_DATE
    if sleeve_name not in SLEEVES:
        print(f"Unknown sleeve: {sleeve_name}. Available: {list(SLEEVES.keys())}")
        return None, None, None

    cfg = SLEEVES[sleeve_name]
    print(f"\n{'='*60}")
    print(f"  Sleeve {sleeve_name}: {cfg['name']}  |  {cfg['cap_tier']}")
    if ANALYSIS_DATE != date.today():
        print(f"  ⏪ BACKTEST MODE — Analysis date: {ANALYSIS_DATE}")
    print(f"{'='*60}")

    # Sleeve D / Liquid — no stocks, but still produce an Excel info sheet
    if not cfg.get("universe"):
        cap = cfg.get("capital", 0)
        mkt = cfg.get("market", "INDIA")
        cur = "₹" if mkt == "INDIA" else "$"
        print(f"\n  Sleeve D — Liquid Buffer")
        print(f"  Capital: {cur}{cap:,.0f}  (earns ~6-7% in Liquid Bees / FD)")
        print(f"  Action: No RS scoring needed. Hold in Liquid Bees ETF.")
        print(f"  In BEAR regime: receives capital released from Sleeves A, B, C.")
        print(f"  In BULL regime: deploy back into Sleeves A, B, C as directed.")
        liquid_regime = {"regime": "N/A", "exposure": 1.0, "vals": {}}
        return liquid_regime, pd.DataFrame(), {
            "cfg":           cfg,
            "all_scores":    pd.DataFrame(),
            "top_df":        pd.DataFrame(),
            "actions":       pd.DataFrame(),
            "regime_info":   liquid_regime,
            "comparison_df": pd.DataFrame(),
            "is_liquid":     True,
        }

    cache    = SimpleCache(CACHE_DIR)
    universe = load_universe(cfg)
    if universe.empty:
        print("  No stocks in universe after exclusion. Check CSV files.")
        return None, pd.DataFrame(), None

    all_scores, regime_info, top_df, close_df = score_today(cfg, universe, cache)

    if top_df.empty:
        print("  No stocks passed the RS peer filter.")
        return regime_info, pd.DataFrame(), None

    exp_frac = regime_info.get("exposure", 1.0)
    today    = close_df.index[-1] if not close_df.empty else pd.Timestamp(ANALYSIS_DATE)
    top_df   = add_weights_to_top(top_df, close_df, today, cfg, exp_frac)

    holdings = load_holdings(sleeve_name)
    print(f"  Holdings on file: {len(holdings)} positions for sleeve {sleeve_name}")

    actions = compute_actions(holdings, top_df, all_scores, cfg, regime_info)
    print_report(sleeve_name, cfg, top_df, actions, regime_info)

    # ── Load previous snapshot for comparison ─────────────────────────────────
    prev_snap      = load_prev_snapshot(sleeve_name, ANALYSIS_DATE, max_lookback_days=90)
    comparison_df  = compare_snapshots(top_df, prev_snap, cfg)
    if not comparison_df.empty:
        n_new  = len(comparison_df[comparison_df["Status"] == "NEW"])
        n_sold = len(comparison_df[comparison_df["Status"] == "REMOVED"])
        n_adj  = len(comparison_df[comparison_df["Action"].str.contains("ADD|REDUCE", na=False)])
        print(f"\n  📊 vs {prev_snap.get('date','?')}: "
              f"{n_new} NEW  |  {n_sold} REMOVED  |  {n_adj} qty adjustments")
    else:
        print("  ℹ  No previous snapshot found (first run or >90 days gap).")

    # ── Save snapshot for future comparison ───────────────────────────────────
    save_snapshot(sleeve_name, top_df, regime_info, ANALYSIS_DATE)

    # Return all data needed for the combined Excel (caller writes it)
    return regime_info, top_df, {
        "cfg":           cfg,
        "all_scores":    all_scores,
        "top_df":        top_df,
        "actions":       actions,
        "regime_info":   regime_info,
        "comparison_df": comparison_df,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  ⓯  INTERACTIVE MENU  (NEW in v2.0)
# ══════════════════════════════════════════════════════════════════════════════

def interactive_menu():
    """Show interactive selection menu. Returns (list_of_sleeves, run_all_flag)."""
    global ANALYSIS_DATE

    os.system("cls" if os.name == "nt" else "clear")

    W = 76
    def box(text="", fill=" "):
        inner = text.center(W - 2)
        return f"║{inner}║"

    border_top = "╔" + "═" * (W - 2) + "╗"
    border_sep = "╠" + "═" * (W - 2) + "╣"
    border_bot = "╚" + "═" * (W - 2) + "╝"

    print(border_top)
    print(box())
    print(box("RS REBALANCING TRACKER  v2.0"))
    print(box("Cap-Tier Separated Momentum System  |  India + USA"))
    print(box())
    print(border_sep)
    print(box("  ── INDIA SLEEVES ──────────────────────────────────────────"))
    print(box("  1.  Sleeve A   — India Large Cap    (Nifty Top 50)  Monthly"))
    print(box("  2.  Sleeve B   — India Mid Cap      (Nifty 51-200)  Fortnightly"))
    print(box("  3.  Sleeve C   — India Small Cap    (Nifty 201-500) Weekly"))
    print(box("  4.  Sleeve D   — India Liquid Bees  (Bear Buffer)"))
    print(border_sep)
    print(box("  ── USA SLEEVES ─────────────────────────────────────────────"))
    print(box("  5.  US_A      — US Mega Cap   (S&P Top 50)          Monthly"))
    print(box("  6.  US_B      — US Large Cap  (S&P 51-200)          Fortnightly"))
    print(box("  7.  US_C      — US Mid Cap    (S&P 201-500)         Weekly"))
    print(box("  8.  US_ETF    — World + Commodity + Defensive ETFs  Monthly"))
    print(border_sep)
    print(box("  ── RUN MULTIPLE ────────────────────────────────────────────"))
    print(box("  9.  All India  (A + B + C + D + overlap check)"))
    print(box("  0.  ALL SLEEVES — India + USA (full portfolio run)"))
    print(border_bot)

    # Choice
    while True:
        choice = input(f"\n  Enter choice (0-9): ").strip()
        if choice in [str(i) for i in range(10)]:
            break
        print("  ⚠  Invalid — please enter a number 0-9.")

    choice_map = {
        "1": ["A"],
        "2": ["B"],
        "3": ["C"],
        "4": ["D"],
        "5": ["US_A"],
        "6": ["US_B"],
        "7": ["US_C"],
        "8": ["US_ETF"],
        "9": ["A", "B", "C", "D"],
        "0": ["A", "B", "C", "D", "US_A", "US_B", "US_C", "US_ETF"],
    }
    sleeves_to_run = choice_map[choice]
    run_all = len(sleeves_to_run) > 1

    # Analysis date
    print(f"\n  ── ANALYSIS DATE ─────────────────────────────────────────────")
    print(f"  Default = today ({date.today().strftime('%Y-%m-%d')})")
    print(f"  For backtesting, enter any past date, e.g.  2024-09-30")
    print(f"  This lets you see last month's list and compare with current.")
    date_input = input(f"\n  Enter date [YYYY-MM-DD] or press Enter for today: ").strip()

    if date_input == "":
        ANALYSIS_DATE = date.today()
    else:
        try:
            parts = date_input.split("-")
            candidate = date(int(parts[0]), int(parts[1]), int(parts[2]))
            if candidate > date.today():
                print("  ⚠  Future date — using today instead.")
                ANALYSIS_DATE = date.today()
            else:
                ANALYSIS_DATE = candidate
        except Exception:
            print("  ⚠  Invalid format — using today.")
            ANALYSIS_DATE = date.today()

    backtest_note = ""
    if ANALYSIS_DATE != date.today():
        backtest_note = f"  ⏪ BACKTEST MODE"

    print(f"\n  {'─'*72}")
    print(f"  Running:       {', '.join(sleeves_to_run)}")
    print(f"  Analysis date: {ANALYSIS_DATE}{backtest_note}")
    print(f"  {'─'*72}\n")

    return sleeves_to_run, run_all


# ══════════════════════════════════════════════════════════════════════════════
#  ⓰  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    global ANALYSIS_DATE

    # ── Legacy CLI mode (backward compatible with v1.x scripts / schedulers) ─
    if len(sys.argv) > 1:
        parser = argparse.ArgumentParser(description="RS Rebalancing Tracker v2.0")
        parser.add_argument("--sleeve", default="A",
                            help="A, B, C, D, US_A, US_B, US_C, US_ETF")
        parser.add_argument("--all",  action="store_true",
                            help="Run all sleeves + overlap check")
        parser.add_argument("--date", default=None,
                            help="Analysis date YYYY-MM-DD  (default: today)")
        args = parser.parse_args()

        if args.date:
            try:
                ANALYSIS_DATE = date.fromisoformat(args.date)
            except Exception:
                print(f"  ⚠  Invalid --date format. Using today.")

        if args.all:
            sleeves_to_run = list(SLEEVES.keys())
            run_all = True
        else:
            sleeves_to_run = [args.sleeve.upper()]
            run_all = False
    else:
        # ── Interactive menu mode (double-click / python script.py) ──────────
        sleeves_to_run, run_all = interactive_menu()

    # ── Run selected sleeves ──────────────────────────────────────────────────
    all_tops      = {}
    all_regime    = {}
    sleeve_data   = {}   # { sleeve_name: data_dict } for combined Excel

    for sname in sleeves_to_run:
        try:
            result = run_sleeve(sname)
            if result is None:
                continue
            rinfo, top, data = result
            if top is not None and not top.empty:
                all_tops[sname] = top
            if rinfo:
                all_regime[sname] = rinfo
            if data:
                sleeve_data[sname] = data
        except Exception as e:
            print(f"  ERROR in sleeve {sname}: {e}")
            import traceback; traceback.print_exc()

    # ── Cross-sleeve overlap check (India only) ───────────────────────────────
    india_tops = {k: v for k, v in all_tops.items() if k in ["A", "B", "C"]}
    if run_all and len(india_tops) >= 2:
        check_universe_overlap(india_tops)

    # ── Portfolio summary (when running multiple) ─────────────────────────────
    if run_all and len(sleeves_to_run) > 1:
        print("\n" + "═" * 70)
        print("  PORTFOLIO SUMMARY — ALL SLEEVES")
        print("═" * 70)
        india_cap = sum(SLEEVES[s].get("capital", 0)
                        for s in ["A", "B", "C", "D"] if s in SLEEVES)
        for sname in sleeves_to_run:
            cfg    = SLEEVES[sname]
            cap    = cfg.get("capital", 0)
            pct    = cap / india_cap * 100 if india_cap and cfg["market"] == "INDIA" else 0
            regime = all_regime.get(sname, {}).get("regime", "N/A")
            n_pos  = len(all_tops.get(sname, pd.DataFrame()))
            mkt    = cfg.get("market", "INDIA")
            cur    = "₹" if mkt == "INDIA" else "$"
            cap_s  = f"{cur}{cap/1e5:.1f}L" if cap > 0 else f"{cur}—"
            pct_s  = f"{pct:.0f}%" if pct > 0 else "—"
            print(f"  {sname:<8} {cfg['name']:<14} {cfg['cap_tier']:<38} "
                  f"{cap_s:<8} {pct_s:<5}  [{regime}]  {n_pos} positions")
        print(f"  {'─'*68}")
        if india_cap > 0:
            print(f"  India total:  ₹{india_cap/1e5:.1f}L")

    # ── Write per-market Excel files (#7): India sleeves → India file, ────────
    #    US sleeves → US file. Every sheet per sleeve is kept
    #    (Actions / RS Rankings / Holdings [+ Comparison]).
    if sleeve_data:
        out_dir  = Path(OUTPUT_DIR); out_dir.mkdir(exist_ok=True)
        date_str = ANALYSIS_DATE.strftime("%Y%m%d")

        # group sleeves by market, preserving insertion order
        by_market = {}
        for sname, data in sleeve_data.items():
            mkt = data["cfg"].get("market", "INDIA")
            by_market.setdefault(mkt, []).append((sname, data))

        for mkt, items in by_market.items():
            outpath = str(out_dir / f"RS_Rebalance_{mkt}_{date_str}.xlsx")
            wb = openpyxl.Workbook()
            wb.remove(wb.active)   # remove default empty sheet
            for sname, data in items:
                try:
                    wb = write_sleeve_sheets(
                        wb, sname,
                        data["cfg"], data["all_scores"], data["top_df"],
                        data["actions"], data["regime_info"],
                        comparison_df=data.get("comparison_df"),
                    )
                except Exception as e:
                    print(f"  ERROR writing sheets for sleeve {sname}: {e}")

            save_combined_excel(wb, outpath)
            print(f"\n  ✅ {mkt} Excel saved → {outpath}")
            print(f"     Sleeves: " + ", ".join(s for s, _ in items))

    if os.name == "nt" and len(sys.argv) == 1:
        input("\n  ✅ Done! Press Enter to close …")


if __name__ == "__main__":
    main()
