"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  EXCEL FORMATTER  v6.0  — market_excel.py                                 ║
║  Shared styling for India & US reports                                     ║
║                                                                            ║
║  v6.0 changes:                                                             ║
║   • Signal_Label coloring — 10 distinct human-readable labels, each with   ║
║     its own color tier (Prime=gold-green, Confirmed=green, RS=light-green, ║
║     Watch=amber, Neutral=grey, Breakdown=red)                              ║
║   • Simplified STOCK_MAIN_COLS — 22 cols vs 30+ before; raw signals       ║
║     only in 🔬 Signal Detail sheet                                         ║
║   • New sheet order: Dashboard > Opportunities > Sectors > Stocks >        ║
║     Global > Sleeves > Breadth > Snapshot > Signal Detail                  ║
║   • "🎯 Opportunities" replaces "🏆 Top Picks - Buy" — cleaner name       ║
║   • Action_Tier coloring extended to Signal_Label for full compatibility   ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""
import numpy as np, pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ─────────────────────────────────────────────────────────────────────────────
#  COLOR PALETTE
# ─────────────────────────────────────────────────────────────────────────────
TAB_COLORS = {
    "📋 Dashboard":        "0A1628",
    "🎯 Opportunities":    "1A6B2E",
    "🏭 Sectors":          "004D40",
    "📊 Stocks":           "0D47A1",
    "🌍 Global":           "1A3C5E",
    "📋 RS Sleeves":       "1A3A5C",
    "📊 Breadth":          "1B5E20",
    "📸 Snapshot":         "0A1628",
    "🔬 Signal Detail":    "37474F",
    # Legacy names kept for backward compatibility
    "📸 Market Snapshot":  "0A1628",
    "🏭 Sector Strength":  "004D40",
    "🔄 Sector Rotation":  "1A237E",
    "🏭 Industry Rotation":"006064",
    "📊 Market Breadth":   "1B5E20",
    "📈 Sector Performance":"4A148C",
    "📊 Stock Strength":   "0D47A1",
    "📊 Signal Detail":    "1A3A5C",
    "🏆 Top Picks - Buy":  "1B5E20",
    "🔴 Top Picks - Sell": "B71C1C",
    "📐 Chart Patterns":   "880E4F",
    "🎯 Trade Setups":     "E65100",
    "📋 RS Sleeve Lists":  "1A3A5C",
    "🌍 Country ETFs":     "1A3C5E",
    "🏅 Commodities":      "4E342E",
}
HDR_COLORS = {
    "📋 Dashboard":        "0D2137",
    "🎯 Opportunities":    "1B5E20",
    "🏭 Sectors":          "00695C",
    "📊 Stocks":           "0D47A1",
    "🌍 Global":           "1A3C5E",
    "📋 RS Sleeves":       "1A3A5C",
    "📊 Breadth":          "2E7D32",
    "📸 Snapshot":         "0D2137",
    "🔬 Signal Detail":    "37474F",
    # Legacy
    "📸 Market Snapshot":  "0D2137",
    "🏭 Sector Strength":  "00695C",
    "🔄 Sector Rotation":  "1A237E",
    "🏭 Industry Rotation":"006064",
    "📊 Market Breadth":   "2E7D32",
    "📈 Sector Performance":"4A148C",
    "📊 Stock Strength":   "0D47A1",
    "📊 Signal Detail":    "1A3A5C",
    "🏆 Top Picks - Buy":  "1B5E20",
    "🔴 Top Picks - Sell": "B71C1C",
    "📐 Chart Patterns":   "880E4F",
    "🎯 Trade Setups":     "BF360C",
    "🌍 Country ETFs":     "1A3C5E",
    "🏅 Commodities":      "4E342E",
}

# ── Signal_Label color map ─────────────────────────────────────────────────────
#  bg_hex, fg_hex, bold
SIGNAL_LABEL_COLORS = {
    "🌟 Triple Confirmed": ("0D3320", "4ADE80", True),
    "🌟 RS30 + Long":      ("0F3B24", "6EE7A0", True),
    "🌟 RS30 + Swing":     ("0F3B24", "6EE7A0", True),
    "🌟 RS30 Leader":      ("14532D", "86EFAC", True),
    "🌟 Long Momentum":    ("14532D", "86EFAC", True),
    "🌟 Prime Setup":      ("166534", "BBF7D0", True),
    "✅ Long Momentum":    ("C8E6C9", "1B5E20", True),
    "✅ Strong RS":        ("DCEDC8", "2E7D32", True),
    "📈 Swing Entry":      ("E3F2FD", "0D47A1", False),
    "📈 RS Leader":        ("E8F5E9", "1B5E20", False),
    "👁 Setup Building":   ("FFF9C4", "795548", False),
    "👁 RS30 Watch":       ("FFF9C4", "5D4037", False),
    "👁 LST Watch":        ("FFF9C4", "5D4037", False),
    "👁 MST Watch":        ("FFF9C4", "5D4037", False),
    "👁 Watch":            ("FFF8E1", "6D4C41", False),
    "⬜ Neutral":          ("F5F5F5", "757575", False),
    "🔴 RS Breakdown":     ("FFCDD2", "B71C1C", True),
}

# ── Legacy signal / action_tier color map ──────────────────────────────────────
SIG_COLORS = {
    "Strong Buy": ("FFFFFF", "006B3C"),
    "Buy":        ("000000", "C8E6C9"),
    "Sell":       ("FFFFFF", "B71C1C"),
    "Neutral":    ("5D4037", "FFF9C4"),
    "WAIT":       ("5D4037", "FFF9C4"),
    "BUY":        ("FFFFFF", "1B5E20"),
    "SELL":       ("FFFFFF", "C62828"),
}
ACTION_TIER_COLORS = {
    "PRIME BUY":     ("FFFFFF", "1A6B2E"),
    "CONFIRMED BUY": ("1B5E20", "C8E6C9"),
    "RS BUY":        ("33691E", "DCEDC8"),
    "WATCH":         ("5D4037", "FFF9C4"),
    "NEUTRAL":       ("888888", "F5F5F5"),
    "AVOID":         ("FFFFFF", "B71C1C"),
}
TREND_MAP = {
    "Strong Bullish":   ("C8E6C9","1B5E20"),
    "Bullish":          ("DCEDC8","33691E"),
    "Neutral":          ("FFF9C4","5D4037"),
    "Bearish":          ("FFCDD2","B71C1C"),
    "Strong Bearish":   ("EF9A9A","7F0000"),
    "Mixed":            ("FFE0B2","E65100"),
    "↑ Bullish":        ("C8E6C9","1B5E20"),
    "↓ Bearish":        ("FFCDD2","B71C1C"),
    "→ Recovering":     ("DCEDC8","33691E"),
    "→ Pulling Back":   ("FFE0B2","E65100"),
    "BULLISH":          ("C8E6C9","1B5E20"),
    "MIXED":            ("FFF9C4","5D4037"),
    "BEARISH":          ("FFCDD2","B71C1C"),
}
ZONE_MAP = {
    "Bullish": ("C8E6C9","1B5E20"),
    "Neutral": ("FFF9C4","5D4037"),
    "Bearish": ("FFCDD2","B71C1C"),
}
THIN = Side(style="thin", color="D0D0D0")

def _F(h):   return PatternFill("solid", fgColor=h)
def _fn(bold=False, color="111111", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def _al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _bd():   return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _w(ws, row, col, val, bold=False, fg="111111", bg=None, size=10, h="center", wrap=False):
    c = ws.cell(row, col, val)
    c.font = _fn(bold=bold, color=fg, size=size)
    c.alignment = _al(h, wrap=wrap)
    c.border = _bd()
    if bg: c.fill = _F(bg)
    return c


# ─────────────────────────────────────────────────────────────────────────────
#  CELL COLORIZER  — handles Signal_Label + all legacy columns
# ─────────────────────────────────────────────────────────────────────────────

def _color_cell(cell, col_name, val):
    col = str(col_name).lower().strip()

    # ── NEW: Signal_Label — primary decision column ───────────────────────────
    if col == "signal_label":
        v = str(val or "")
        if v in SIGNAL_LABEL_COLORS:
            bg, fg, bold = SIGNAL_LABEL_COLORS[v]
            cell.fill = _F(bg)
            cell.font = _fn(bold=bold, color=fg)
        return

    # ── Action_Tier (legacy + v5.3) ───────────────────────────────────────────
    if col == "action_tier":
        v = str(val or "")
        if v in ACTION_TIER_COLORS:
            fg, bg = ACTION_TIER_COLORS[v]
            cell.fill = _F(bg)
            cell.font = _fn(bold=(v in ("PRIME BUY","CONFIRMED BUY","AVOID")), color=fg)
        return

    # ── Sec_Gated ─────────────────────────────────────────────────────────────
    if col == "sec_gated":
        v = str(val or "")
        if v == "✓": cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
        elif v == "✗": cell.fill = _F("F5F5F5"); cell.font = _fn(color="AAAAAA")
        return

    # ── Signal / Enhanced / Action columns ────────────────────────────────────
    if col in ("signal", "enhanced", "sec_signal", "action", "rs_signal"):
        v = str(val or "")
        if v in SIG_COLORS:
            fg, bg = SIG_COLORS[v]
            cell.fill = _F(bg); cell.font = _fn(bold=True, color=fg)
        return

    # ── MST / LST / RS30 ──────────────────────────────────────────────────────
    if col in ("mst_signal", "lst_signal", "rs30_signal"):
        v = str(val or "")
        if v == "Buy":
            cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
        elif v == "Watch":
            cell.fill = _F("FFF9C4"); cell.font = _fn(bold=True, color="5D4037")
        elif v == "Neutral":
            cell.fill = _F("F5F5F5"); cell.font = _fn(color="888888")
        return

    # ── Supertrend ────────────────────────────────────────────────────────────
    if col == "supertrend":
        v = str(val or "")
        if v == "Buy":   cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
        elif v == "Sell": cell.fill = _F("FFCDD2"); cell.font = _fn(bold=True, color="B71C1C")
        elif v == "N/A":  cell.fill = _F("F5F5F5"); cell.font = _fn(color="AAAAAA", italic=True)
        return

    # ── Trend ─────────────────────────────────────────────────────────────────
    if "trend" in col:
        for k, (bg, fg) in TREND_MAP.items():
            if k.lower() in str(val or "").lower():
                cell.fill = _F(bg); cell.font = _fn(bold=True, color=fg); return
        return

    # ── Zone ──────────────────────────────────────────────────────────────────
    if "zone" in col:
        v = str(val or "")
        if v in ZONE_MAP:
            bg, fg = ZONE_MAP[v]; cell.fill = _F(bg); cell.font = _fn(bold=True, color=fg)
        return

    # ── Tick ✓/✗ ──────────────────────────────────────────────────────────────
    if col.startswith("abv_") or "beats" in col or col in ("breakout_up", "sec_gated", "w_ema10_gtema30"):
        v = str(val or "")
        if v == "✓": cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
        elif v == "✗": cell.fill = _F("FFCDD2"); cell.font = _fn(color="B71C1C")
        return

    # ── Chart pattern ─────────────────────────────────────────────────────────
    if col == "chart_pattern":
        v = str(val or "")
        if v.startswith("🟢"): cell.fill = _F("E3F2FD"); cell.font = _fn(bold=True, color="1565C0")
        elif v.startswith("🔴"): cell.fill = _F("FFEBEE"); cell.font = _fn(bold=True, color="C62828")
        return

    # ── RSI ───────────────────────────────────────────────────────────────────
    if col in ("rsi_14", "rsi", "w_rsi", "m_rsi"):
        if isinstance(val, (int, float)) and not np.isnan(val):
            if val >= 70:   cell.fill = _F("FFCDD2"); cell.font = _fn(bold=True, color="B71C1C")
            elif val >= 60: cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
            elif val >= 50: cell.fill = _F("F1F8E9"); cell.font = _fn(color="33691E")
            elif val < 30:  cell.fill = _F("FFCDD2"); cell.font = _fn(color="B71C1C")
        return

    # ── SMA Score ─────────────────────────────────────────────────────────────
    if col == "sma_score":
        pal = {4:("C8E6C9","1B5E20"), 3:("DCEDC8","33691E"),
               2:("FFF9C4","5D4037"), 1:("FFE0B2","E65100"), 0:("FFCDD2","B71C1C")}
        try:
            v = int(float(val))
            if v in pal: bg, fg = pal[v]; cell.fill = _F(bg); cell.font = _fn(bold=(v==4), color=fg)
        except: pass
        return

    # ── SL Grade ──────────────────────────────────────────────────────────────
    if col == "sl_grade":
        pal = {"A":("C8E6C9","1B5E20"), "B":("DCEDC8","33691E"),
               "C":("FFF9C4","5D4037"), "D":("FFE0B2","E65100"), "F":("FFCDD2","B71C1C")}
        v = str(val or "")
        if v in pal: bg, fg = pal[v]; cell.fill = _F(bg); cell.font = _fn(bold=(v=="A"), color=fg)
        return

    # ── SL % (inverted — lower = greener) ────────────────────────────────────
    if col in ("sl_buy%", "sl%", "sl_sell%", "sl_buy_pct", "sl_sell_pct"):
        if isinstance(val, (int, float)) and not np.isnan(val):
            if   val <= 3:  cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
            elif val <= 5:  cell.fill = _F("DCEDC8"); cell.font = _fn(color="33691E")
            elif val <= 8:  cell.fill = _F("FFF9C4"); cell.font = _fn(color="5D4037")
            elif val <= 12: cell.fill = _F("FFE0B2"); cell.font = _fn(color="E65100")
            else:           cell.fill = _F("FFCDD2"); cell.font = _fn(color="B71C1C")
        return

    # ── RR Ratio ──────────────────────────────────────────────────────────────
    if col in ("rr_t1", "rr_t2"):
        if isinstance(val, (int, float)) and not np.isnan(val):
            if   val >= 3: cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
            elif val >= 2: cell.fill = _F("FFF9C4"); cell.font = _fn(color="5D4037")
            else:          cell.fill = _F("FFCDD2"); cell.font = _fn(color="B71C1C")
        return

    # ── TV Symbol ─────────────────────────────────────────────────────────────
    if col == "tv_symbol":
        cell.fill = _F("E8F5E9"); cell.font = _fn(bold=True, color="1B5E20", size=9)
        return

    # ── Action (BUY/SELL/WAIT) ────────────────────────────────────────────────
    if col == "action":
        v = str(val or "")
        if v == "BUY":   cell.fill = _F("1B5E20"); cell.font = _fn(bold=True, color="FFFFFF")
        elif v == "SELL": cell.fill = _F("B71C1C"); cell.font = _fn(bold=True, color="FFFFFF")
        elif v == "WAIT": cell.fill = _F("FFF9C4"); cell.font = _fn(bold=True, color="5D4037")
        return

    # ── Breadth scores ────────────────────────────────────────────────────────
    if col in ("rs22%","rs55%","rsi50%","abvsma20%","abvsma50%","abvsma100%","abvsma200%",
               "1m_score","3m_score","6m_score"):
        if isinstance(val, (int, float)) and not np.isnan(val):
            if val >= 60: cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
            elif val >= 40: cell.fill = _F("FFF9C4"); cell.font = _fn(color="5D4037")
            else: cell.fill = _F("FFCDD2"); cell.font = _fn(color="B71C1C")
        return

    # ── ATR / Weight columns ──────────────────────────────────────────────────
    if col in ("atr_wt%", "equal_wt%"):
        if isinstance(val, (int, float)) and not np.isnan(val):
            if val >= 10: cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
            elif val >= 5: cell.fill = _F("F1F8E9"); cell.font = _fn(color="33691E")
            else:          cell.fill = _F("FFF9C4"); cell.font = _fn(color="5D4037")
        return

    if col == "daily_std%":
        if isinstance(val, (int, float)) and not np.isnan(val):
            if   val <= 1.5: cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
            elif val <= 2.5: cell.fill = _F("FFF9C4"); cell.font = _fn(color="5D4037")
            else:            cell.fill = _F("FFCDD2"); cell.font = _fn(color="B71C1C")
        return

    if col == "avg_turnover":
        if isinstance(val, (int, float)) and not np.isnan(val):
            if   val >= 100: cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
            elif val >= 20:  cell.fill = _F("F1F8E9"); cell.font = _fn(color="33691E")
            elif val >= 5:   cell.fill = _F("FFF9C4"); cell.font = _fn(color="5D4037")
            else:            cell.fill = _F("FFCDD2"); cell.font = _fn(color="B71C1C")
        return

    # ── D/E ratio ─────────────────────────────────────────────────────────────
    if col in ("d/e", "de"):
        if isinstance(val, (int, float)) and not np.isnan(val):
            if val < 0.5: cell.fill = _F("C8E6C9"); cell.font = _fn(color="1B5E20")
            elif val < 1: cell.fill = _F("FFF9C4"); cell.font = _fn(color="5D4037")
            elif val >= 2: cell.fill = _F("FFCDD2"); cell.font = _fn(color="B71C1C")
        return

    # ── Signal_Type (trade setups) ────────────────────────────────────────────
    if col in ("signal_type", "setup_desc"):
        v = str(val or "")
        if any(x in v for x in ["RS30","LST","MST"]):
            cell.fill = _F("E3F2FD"); cell.font = _fn(bold=True, color="1565C0")
        elif "Strong" in v:
            cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="006B3C")
        elif "Sell" in v or "🔴" in v:
            cell.fill = _F("FFCDD2"); cell.font = _fn(bold=True, color="B71C1C")
        elif "Watch" in v or "⏳" in v:
            cell.fill = _F("FFF9C4"); cell.font = _fn(color="5D4037")
        return

    # ── Generic % columns — positive=green, negative=red ─────────────────────
    pct_set = {
        "chg_1d%","chg_5d%","avg_chg_1d%","avg_chg_5d%",
        "rs_22d%","rs_55d%","rs_120d%","rs_252d%",
        "1m%","3m%","6m%","12m%","ytd%","rs_1m%","rs_3m%","rs_6m%","rs_12m%",
        "from_52w_high%","rs_score","total_score",
        "sales_qoq%","sales_yoy%","pat_qoq%","pat_yoy%","margin%","roe%",
        "sec_rs22d%","sec_rs55d%","sec_rs120d%","sec_rs252d%",
        "rs_22d_idx%","rs_55d_idx%","rs_120d_idx%","rs_252d_idx%",
        "rs_22d_sec%","rs_55d_sec%","rs_120d_sec%","rs_252d_sec%",
        "ret_22d%","ret_55d%","ret_120d%","ret_252d%",
        "w_rs21%","w_rs30%","m_rs12%","fin_score","rs22%","rs55%",
    }
    if col in pct_set or (col.endswith("%") and col not in ("sl_buy%","sl_sell%","sl%")):
        if isinstance(val, (int, float)) and not np.isnan(val):
            bold = abs(val) > 5
            if val > 0: cell.fill = _F("C8E6C9"); cell.font = _fn(bold=bold, color="1B5E20")
            elif val < 0: cell.fill = _F("FFCDD2"); cell.font = _fn(bold=bold, color="B71C1C")
        return


# ─────────────────────────────────────────────────────────────────────────────
#  CORE SHEET WRITER
#  Row 1 = Title | Row 2 = Info | Row 3 = HEADER | Row 4+ = Data
# ─────────────────────────────────────────────────────────────────────────────

_LEFT_COLS = {
    'symbol','tv_symbol','company','company name','name','sector','industry',
    'chart_pattern','notes','setup_desc','signal_type','strategy','trend',
    'signal_label',   # ← new — left align for readability
}

def write_sheet(ws, df, sheet_name, title="", freeze_row=4):
    if df is None or df.empty:
        ws.cell(1, 1, "No data available for this section."); return

    hdr_bg = HDR_COLORS.get(sheet_name, "0D2137")
    n_cols = len(df.columns)

    # Row 1: Title
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        c = ws.cell(1, 1, title)
        c.fill = _F(hdr_bg); c.font = _fn(bold=True, color="FFFFFF", size=13)
        c.alignment = _al("left"); ws.row_dimensions[1].height = 28

    # Row 2: Info
    ts = datetime.now().strftime("%d-%b-%Y %H:%M")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(n_cols, 10))
    info = ws.cell(2, 1, f"Generated: {ts}  |  {len(df)} rows")
    info.font = _fn(italic=True, color="777777", size=9)
    info.alignment = _al("left"); ws.row_dimensions[2].height = 14

    # Row 3: Header
    headers = list(df.columns)
    for j, h in enumerate(headers, 1):
        c = ws.cell(3, j, h)
        c.fill = _F(hdr_bg); c.font = _fn(bold=True, color="FFFFFF", size=10)
        c.alignment = _al("center", wrap=True); c.border = _bd()
    ws.row_dimensions[3].height = 24
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"
    ws.freeze_panes = f"A{freeze_row}"

    # Rows 4+: Data
    for i, (_, row_data) in enumerate(df.iterrows()):
        r = i + 4
        alt_bg = "F7F9FC" if i % 2 == 0 else "FFFFFF"
        for j, col in enumerate(headers, 1):
            val = row_data[col]
            if val is None or (isinstance(val, float) and np.isnan(val)): val = "—"
            c = ws.cell(r, j, val)
            c.alignment = _al("left" if col.lower() in _LEFT_COLS else "center")
            c.border = _bd(); c.fill = _F(alt_bg); c.font = _fn(size=10)
            _color_cell(c, col, val)
        ws.row_dimensions[r].height = 17

    # Column widths
    ws.column_dimensions["A"].width = 6
    for col_obj in ws.iter_cols(min_row=3, max_row=3):
        j = col_obj[0].column
        ltr = get_column_letter(j)
        hv  = str(col_obj[0].value or "").lower()
        max_data = max(
            (len(str(ws.cell(r, j).value or "")) for r in range(4, ws.max_row + 1)),
            default=8
        )
        if hv in ("rank","sec_rank","valid","adv","dec","h_day","sma_score","fin_score"):
            ws.column_dimensions[ltr].width = 7
        elif hv == "signal_label":    ws.column_dimensions[ltr].width = 22
        elif hv == "action_tier":     ws.column_dimensions[ltr].width = 16
        elif hv == "sec_gated":       ws.column_dimensions[ltr].width = 10
        elif hv == "tv_symbol":       ws.column_dimensions[ltr].width = 22
        elif hv == "symbol":          ws.column_dimensions[ltr].width = 12
        elif hv in ("company","company name","name"): ws.column_dimensions[ltr].width = 28
        elif hv in ("sector","industry"):             ws.column_dimensions[ltr].width = 18
        elif hv == "chart_pattern":   ws.column_dimensions[ltr].width = 24
        elif hv in ("setup_type","notes","trend"):    ws.column_dimensions[ltr].width = min(max_data + 3, 28)
        else:                         ws.column_dimensions[ltr].width = min(max_data + 2, 16)

    ws.sheet_view.showGridLines = False


# ─────────────────────────────────────────────────────────────────────────────
#  OPPORTUNITIES SHEET (replaces Top Picks Buy)
#  Sector-grouped with colored dividers. Signal_Label prominently displayed.
# ─────────────────────────────────────────────────────────────────────────────

def write_opportunities_sheet(ws, df, sheet_name, title="", is_sell=False, primary_rs=55):
    if df is None or df.empty:
        ws.cell(1, 1, "No data available."); return
    if "Message" in df.columns:
        ws.cell(1, 1, df["Message"].iloc[0]); return

    hdr_bg = HDR_COLORS.get(sheet_name, "1B5E20")
    display_cols = [c for c in df.columns if not c.startswith("_")]
    n_cols = len(display_cols)

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, title)
    c.fill = _F(hdr_bg); c.font = _fn(bold=True, color="FFFFFF", size=13)
    c.alignment = _al("left"); ws.row_dimensions[1].height = 28

    # Row 2: Info
    ts = datetime.now().strftime("%d-%b-%Y %H:%M")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(n_cols, 10))
    direction = "strongest→weakest" if not is_sell else "weakest→strongest"
    ws.cell(2, 1, f"Generated: {ts}  |  Sectors ranked {direction} by RS_{primary_rs}d%  |  Signal_Label = unified decision signal")
    ws.cell(2, 1).font = _fn(italic=True, color="777777", size=9)
    ws.cell(2, 1).alignment = _al("left"); ws.row_dimensions[2].height = 14

    # Row 3: Header
    for j, h in enumerate(display_cols, 1):
        c = ws.cell(3, j, h)
        c.fill = _F(hdr_bg); c.font = _fn(bold=True, color="FFFFFF", size=10)
        c.alignment = _al("center", wrap=True); c.border = _bd()
    ws.row_dimensions[3].height = 24
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"
    ws.freeze_panes = "A4"

    # Detect sector RS column
    sec_rs_col = f"Sec_RS{primary_rs}d%"
    if sec_rs_col not in display_cols:
        sec_rs_col = next((c for c in display_cols if c.startswith("Sec_RS") and c.endswith("%")), None)

    # Data with sector-group divider rows
    r = 4; prev_sec = None
    for i, (_, row_data) in enumerate(df.iterrows()):
        sec = row_data.get("Sector", "")
        if sec != prev_sec:
            sec_rs  = row_data.get(sec_rs_col, np.nan) if sec_rs_col else np.nan
            sec_sig = row_data.get("Sec_Signal", "")
            sec_rk  = row_data.get("Sec_Rank", "")
            if not (isinstance(sec_rs, float) and np.isnan(sec_rs)):
                div_text = f"  ▸  #{sec_rk}  {sec}   {sec_sig}   RS_{primary_rs}d: {sec_rs:+.1f}%"
            else:
                div_text = f"  ▸  #{sec_rk}  {sec}   {sec_sig}"
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
            dc = ws.cell(r, 1, div_text)
            div_bg = "C8E6C9" if sec_sig == "Buy" else ("FFCDD2" if sec_sig == "Sell" else "FFF9C4")
            div_fg = "1B5E20" if sec_sig == "Buy" else ("B71C1C" if sec_sig == "Sell" else "5D4037")
            dc.fill = _F(div_bg); dc.font = _fn(bold=True, color=div_fg, size=10)
            dc.alignment = _al("left"); dc.border = _bd()
            ws.row_dimensions[r].height = 20; r += 1; prev_sec = sec

        alt_bg = "F0FFF0" if not is_sell else "FFF5F5"
        alt_bg = alt_bg if i % 2 == 0 else "FFFFFF"
        for j, col in enumerate(display_cols, 1):
            val = row_data.get(col, "")
            if val is None or (isinstance(val, float) and np.isnan(val)): val = "—"
            c = ws.cell(r, j, val)
            c.alignment = _al("left" if col.lower() in _LEFT_COLS else "center")
            c.border = _bd(); c.fill = _F(alt_bg); c.font = _fn(size=10)
            _color_cell(c, col, val)
        ws.row_dimensions[r].height = 17; r += 1

    # Column widths
    ws.column_dimensions["A"].width = 6
    for col_obj in ws.iter_cols(min_row=3, max_row=3):
        j = col_obj[0].column; ltr = get_column_letter(j); h = str(col_obj[0].value or "").lower()
        if h in ("rank","sec_rank"):      ws.column_dimensions[ltr].width = 7
        elif h == "signal_label":         ws.column_dimensions[ltr].width = 22
        elif h == "action_tier":          ws.column_dimensions[ltr].width = 16
        elif h == "sec_gated":            ws.column_dimensions[ltr].width = 10
        elif h == "tv_symbol":            ws.column_dimensions[ltr].width = 22
        elif h == "symbol":               ws.column_dimensions[ltr].width = 12
        elif h in ("company","company name"): ws.column_dimensions[ltr].width = 28
        elif h in ("sector","industry"):  ws.column_dimensions[ltr].width = 18
        elif h == "chart_pattern":        ws.column_dimensions[ltr].width = 24
        elif h == "trend":                ws.column_dimensions[ltr].width = 16
        else:                             ws.column_dimensions[ltr].width = min(len(str(col_obj[0].value or "")) + 3, 16)
    ws.sheet_view.showGridLines = False


# keep alias for backward compatibility
write_top_picks_sheet = write_opportunities_sheet


# ─────────────────────────────────────────────────────────────────────────────
#  RS SLEEVE SHEET WRITER
# ─────────────────────────────────────────────────────────────────────────────

_SLEEVE_DIVIDER_BG  = "1A3A5C"
_SLEEVE_LEGEND_BG   = "263238"
_SLEEVE_A_BG        = "E3F2FD"
_SLEEVE_B_BG        = "E8F5E9"
_SLEEVE_C_BG        = "FFF8E1"
_SLEEVE_USA_BG      = "F3E5F5"


def write_rs_sleeve_sheet(ws, df, market="INDIA"):
    ws.sheet_view.showGridLines = False
    title = (
        f"📋  RS SLEEVE LISTS  [{market}]  —  Smallcase / MF-style Momentum Portfolios"
        f"  |  A = Large Cap (Monthly)  ·  B = Growth (Fortnightly)  ·  C = Aggressive (Weekly)"
    )
    n_cols = max(len(df.columns), 15)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(1, 1, title)
    tc.fill = _F(_SLEEVE_DIVIDER_BG); tc.font = _fn(bold=True, color="00E5FF", size=12)
    tc.alignment = _al("left"); ws.row_dimensions[1].height = 26

    headers = list(df.columns)
    hdr_display = {
        "Rank":"Rank","Symbol":"Symbol","Company":"Company / Description",
        "Sector":"Sector","Industry":"Industry","Price":"Price / Detail",
        "Sleeve_RS":"Sleeve RS","Signal_Label":"Signal",
        "RS_22d_Idx%":"RS 22d%","RS_55d_Idx%":"RS 55d%",
        "RS_120d_Idx%":"RS 120d%","RS_252d_Idx%":"RS 252d%",
        "Signal":"Signal","RSI_14":"RSI","Supertrend":"ST",
        "SL_Buy%":"SL%","SL_Grade":"SL Gr","SL_Buy_Price":"SL Price",
        "MST_Signal":"MST","LST_Signal":"LST","RS30_Signal":"RS30",
        "Sales_YoY%":"Sales YoY%","PAT_YoY%":"PAT YoY%","ROE%":"ROE%",
        "Mkt_Cap_B":"MCap B","Chart_Pattern":"Pattern",
        "Equal_Wt%":"Eq Wt%","ATR_Wt%":"ATR Wt%","Avg_Turnover":"Turnover",
        "Daily_Std%":"Std%",
    }
    for j, h in enumerate(headers, 1):
        c = ws.cell(2, j, hdr_display.get(h, h))
        c.fill = _F(_SLEEVE_DIVIDER_BG); c.font = _fn(bold=True, color="FFFFFF", size=10)
        c.alignment = _al("center", wrap=True); c.border = _bd()
    ws.row_dimensions[2].height = 22; ws.freeze_panes = "A3"

    cur_sleeve = "A"; cur_bg = _SLEEVE_A_BG
    _sleeve_bg_map = {
        "A":_SLEEVE_A_BG,"B":_SLEEVE_B_BG,"C":_SLEEVE_C_BG,
        "US_A":_SLEEVE_USA_BG,"US_B":_SLEEVE_B_BG,"US_C":_SLEEVE_C_BG,
    }
    _REGIME_BG = {"BULL":"1B5E20","CAUTION":"E65100","BEAR":"B71C1C"}

    for i, (_, row_data) in enumerate(df.iterrows()):
        r = i + 3
        rank_val = str(row_data.get("Rank","") or "")
        is_divider = rank_val.startswith("━━━")
        is_blank   = all(str(v or "").strip() == "" for v in row_data.values)
        if is_blank: ws.row_dimensions[r].height = 7; continue

        if is_divider:
            for key in _sleeve_bg_map:
                if f"SLEEVE {key}" in rank_val:
                    cur_sleeve = key; cur_bg = _sleeve_bg_map.get(key, _SLEEVE_A_BG); break
            is_regime = "MARKET REGIME" in rank_val.upper()
            is_legend  = "METHOD" in rank_val.upper()
            if is_regime:
                sym_val = str(row_data.get("Symbol","") or "")
                div_bg = _REGIME_BG["BEAR"] if ("BEAR" in sym_val and "CAUTION" not in sym_val) \
                    else (_REGIME_BG["CAUTION"] if "CAUTION" in sym_val else _REGIME_BG["BULL"])
            elif is_legend: div_bg = _SLEEVE_LEGEND_BG
            else:           div_bg = _SLEEVE_DIVIDER_BG
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
            parts = [str(row_data.get(h,"") or "") for h in
                     ["Rank","Symbol","Company","Sector","Industry","Price","Sleeve_RS"]
                     if str(row_data.get(h,"") or "").strip()]
            dc = ws.cell(r, 1, "   |   ".join(parts[:6]))
            dc.fill = _F(div_bg); dc.font = _fn(bold=True, color="00E5FF" if is_legend else "FFFFFF", size=10)
            dc.alignment = _al("left"); dc.border = _bd()
            ws.row_dimensions[r].height = 22 if is_regime else 20; continue

        alt_bg = cur_bg if i % 2 == 0 else "FFFFFF"
        _LEFT = {'symbol','company','sector','industry','chart_pattern','signal','signal_label',
                 'enhanced','mst_signal','lst_signal','rs30_signal','supertrend'}
        for j, col in enumerate(headers, 1):
            val = row_data.get(col, "")
            if val is None or (isinstance(val, float) and np.isnan(val)): val = "—"
            c = ws.cell(r, j, val)
            c.alignment = _al("left" if col.lower() in _LEFT else "center")
            c.border = _bd(); c.fill = _F(alt_bg); c.font = _fn(size=10)
            _color_cell(c, col, val)
        ws.row_dimensions[r].height = 17

    col_widths = {
        "Rank":6,"Symbol":12,"Company":28,"Sector":16,"Industry":18,
        "Price":10,"Sleeve_RS":12,"Signal_Label":22,
        "RS_22d_Idx%":9,"RS_55d_Idx%":9,"RS_120d_Idx%":10,"RS_252d_Idx%":10,
        "Avg_Turnover":12,"Daily_Std%":10,"Equal_Wt%":9,"ATR_Wt%":9,
        "Signal":10,"RSI_14":7,"Supertrend":8,"SL_Buy%":8,"SL_Grade":7,
        "SL_Buy_Price":10,"MST_Signal":7,"LST_Signal":7,"RS30_Signal":7,
        "Sales_YoY%":9,"PAT_YoY%":9,"ROE%":7,"Mkt_Cap_B":9,"Chart_Pattern":22,
    }
    for j, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = col_widths.get(col, 12)


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN WORKBOOK BUILDER  v6.0 — redesigned sheet order
#
#  Sheet order (decision-focused):
#   0  📋 Dashboard      — run summary + signal guide + TV watchlists
#   1  🎯 Opportunities  — prime+confirmed buy stocks (renamed Top Picks Buy)
#   2  🔴 Sell Alerts    — sell signals
#   3  🏭 Sectors        — sector strength (primary decision tree)
#   4  🔄 Rotation       — sector + industry rotation breadth
#   5  📊 Stocks         — main stock table (simplified 22-col view)
#   6  🎯 Trade Setups   — BUY/SELL/WAIT with SL+TP+RR
#   7  🌍 Global         — country ETFs + commodities
#   8  📋 RS Sleeves     — smallcase-style momentum portfolios
#   9  📊 Breadth        — market breadth + advance/decline
#   10 📈 Sector Perf.  — 1M/3M/6M/YTD returns
#   11 📸 Snapshot       — index + commodity + forex snapshot
#   12 📐 Patterns       — chart patterns (daily + weekly)
#   13 🔬 Signal Detail  — ALL raw signal columns (reference only)
# ─────────────────────────────────────────────────────────────────────────────

# Simplified columns for main Stock sheet (22 cols — clear, decision-focused)
STOCK_MAIN_COLS = [
    "Symbol", "Company", "Sector",
    "Price", "Chg_1D%",
    "Signal_Label",       # ← PRIMARY: unified human-readable decision
    "Sec_Gated",          # ← sector quality gate
    "RS_22d_Idx%", "RS_55d_Idx%",
    "RSI_14", "Trend", "SMA_Score",
    "SL_Buy%", "SL_Grade", "SL_Buy_Price",
    "Total_Score", "Fin_Score",
    "Sales_YoY%", "PAT_YoY%", "ROE%", "D/E", "EPS",
    "Chart_Pattern",
]

STOCK_MAIN_COLS_V6 = [
    # Identity
    "Symbol", "Company", "Sector", "Industry",
    # Price + momentum
    "Price", "Chg_1D%", "Chg_5D%",
    # PRIMARY DECISION COLUMN (new unified label)
    "Signal_Label",
    # Sector gate quality flag
    "Sec_Gated",
    # Core RS metrics
    "RS_22d_Idx%", "RS_55d_Idx%",
    # Technical health
    "RSI_14", "Trend", "SMA_Score",
    # Risk management
    "SL_Buy%", "SL_Grade", "SL_Buy_Price",
    # Score for ranking
    "Total_Score", "Fin_Score",
    # Fundamentals (key 4 only)
    "Sales_YoY%", "PAT_YoY%", "ROE%", "D/E",
    # Size + pattern
    "Mkt_Cap_B", "Chart_Pattern",
]

# Columns to include in Opportunities / Top Picks sheet (minimal, decision-focused)
OPPORTUNITIES_COLS_V6 = [
    "Rank", "Sec_Rank", "Sector", "Sec_Signal",
    "Signal_Label",
    "Symbol", "Company", "Price", "Chg_1D%",
    "RS_22d_Idx%", "RSI_14",
    "SL_Buy%", "SL_Grade", "SL_Buy_Price",
    "Total_Score", "Fin_Score",
    "Sales_YoY%", "PAT_YoY%", "ROE%",
    "Chart_Pattern",
]

# Signal_Label sort order for display (best signals first)
SIGNAL_LABEL_ORDER = {
    "🌟 Triple Confirmed": 0,
    "🌟 RS30 + Long":      1,
    "🌟 RS30 + Swing":     2,
    "🌟 RS30 Leader":      3,
    "🌟 Long Momentum":    4,
    "🌟 Prime Setup":      5,
    "✅ Long Momentum":    6,
    "✅ Strong RS":        7,
    "📈 Swing Entry":      8,
    "📈 RS Leader":        9,
    "👁 Setup Building":   10,
    "👁 RS30 Watch":       11,
    "👁 LST Watch":        12,
    "👁 MST Watch":        13,
    "👁 Watch":            14,
    "⬜ Neutral":          15,
    "🔴 RS Breakdown":     16,
}
def build_workbook(market, snapshot_df, sector_str_df, sector_rot_df,
                   industry_rot_df, breadth_df, sector_perf_df, stock_str_df,
                   top_buy_df, top_sell_df, chart_pat_df, trade_df, output_path,
                   dashboard_df=None, sleeve_df=None, primary_rs=55,
                   country_etf_df=None, commodity_df=None):

    wb = Workbook(); wb.remove(wb.active)
    run_time = datetime.now().strftime("%d %b %Y  %H:%M")

    # ── 1. Dashboard (official starting point) ────────────────────────────────
    ws = wb.create_sheet("📋 Dashboard")
    if dashboard_df is not None and not dashboard_df.empty:
        write_dashboard_sheet(ws, dashboard_df, market)
    else:
        ws.cell(1, 1, "Dashboard — run market analysis to populate.")
    ws.sheet_view.showGridLines = False

    # ── 2. Snapshots ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("📸 Snapshot")
    write_sheet(ws, snapshot_df, "📸 Market Snapshot",
                f"📸  MARKET SNAPSHOT  [{market}]  —  {run_time}")

    # ── 3. Market Global (Country ETFs + Commodities) ─────────────────────────
    combined_global = _combine_global(country_etf_df, commodity_df)
    ws = wb.create_sheet("🌍 Global")
    write_sheet(ws, combined_global, "🌍 Country ETFs",
                f"🌍  GLOBAL — Country ETFs (RS vs SPY) + Commodities (RS vs GLD)"
                f"  |  Ranked by RS_{primary_rs}d%")

    # ── 4. Sector Rotation (ISOLATED — kept separate from Industry) ───────────
    ws = wb.create_sheet("🔄 Sector Rotation")
    write_sheet(ws, sector_rot_df, "🔄 Sector Rotation",
                "🔄  SECTOR ROTATION  —  Breadth, RS%, Zone Scores")

    # ── 5. Industry Rotation (ISOLATED — separate sheet from Sector) ──────────
    ws = wb.create_sheet("🏭 Industry Rotation")
    write_sheet(ws, industry_rot_df, "🏭 Industry Rotation",
                "🏭  INDUSTRY ROTATION  —  Breadth, RS%, Zone Scores")

    # ── 6. Market Breadth ─────────────────────────────────────────────────────
    ws = wb.create_sheet("📊 Breadth")
    write_sheet(ws, breadth_df, "📊 Market Breadth",
                f"📊  MARKET BREADTH  [{market}]  —  Adv/Dec + % Above MAs")

    # ── 7. Chart Pattern ──────────────────────────────────────────────────────
    ws = wb.create_sheet("📐 Patterns")
    write_sheet(ws, chart_pat_df, "📐 Chart Patterns",
                "📐  CHART PATTERNS  —  Weekly + Daily  |  Entry / Stop / Target / RR")

    # ── 8. Opportunities (parent buy tracker = Buy Alerts) ────────────────────
    ws = wb.create_sheet("🎯 Opportunities")
    write_opportunities_sheet(
        ws, top_buy_df, "🎯 Opportunities",
        "🎯  OPPORTUNITIES — Best Buy Setups by Sector  "
        "|  Signal_Label = unified decision  |  Sector: strongest→weakest",
        is_sell=False, primary_rs=primary_rs,
    )

    # ── 9/10. Sell Alerts (separated cleanly from Buy) ────────────────────────
    ws = wb.create_sheet("🔴 Sell Alerts")
    write_opportunities_sheet(
        ws, top_sell_df, "🔴 Top Picks - Sell",
        "🔴  SELL ALERTS — RS Breakdown Stocks  |  Weakest Sector → Strongest",
        is_sell=True, primary_rs=primary_rs,
    )

    # ── 11. Trade Setup ───────────────────────────────────────────────────────
    ws = wb.create_sheet("🎯 Trade Setups")
    write_sheet(ws, trade_df, "🎯 Trade Setups",
                "🎯  TRADE SETUPS  —  BUY/SELL/WAIT  |  SL + TP1 + TP2 + R:R")

    # ── 12. Sleeves (Portfolio allocations) ───────────────────────────────────
    if sleeve_df is not None and not sleeve_df.empty:
        ws = wb.create_sheet("📋 RS Sleeves")
        write_rs_sleeve_sheet(ws, sleeve_df, market)

    # ── APPENDIX — sheets kept from the prior workbook (not in the 12-sheet ────
    #    target order). Non-destructive: data preserved. Move/remove on request.
    # Sector Strength
    ws = wb.create_sheet("🏭 Sectors")
    write_sheet(ws, sector_str_df, "🏭 Sector Strength",
                f"🏭  SECTOR STRENGTH  —  Ranked by RS_{primary_rs}d% vs Benchmark")

    # Sector Performance
    ws = wb.create_sheet("📈 Sector Perf")
    write_sheet(ws, sector_perf_df, "📈 Sector Performance",
                "📈  SECTOR PERFORMANCE  —  1M / 3M / 6M / YTD Returns")

    # Stocks (simplified main view)
    ws = wb.create_sheet("📊 Stocks")
    if not stock_str_df.empty:
        main_cols = [c for c in STOCK_MAIN_COLS if c in stock_str_df.columns]
        stock_main_df = stock_str_df[main_cols].copy()
    else:
        stock_main_df = stock_str_df
    write_sheet(ws, stock_main_df, "📊 Stock Strength",
                "📊  STOCKS  v6.0  —  Signal_Label  |  See 🔬 Signal Detail for all raw columns")
    # RS colour scale on RS_22d_Idx% column
    if not stock_main_df.empty and "RS_22d_Idx%" in stock_main_df.columns:
        idx_c = list(stock_main_df.columns).index("RS_22d_Idx%") + 1
        ltr = get_column_letter(idx_c)
        ws.conditional_formatting.add(
            f"{ltr}4:{ltr}{len(stock_main_df)+3}",
            ColorScaleRule(start_type="min", start_color="FF4444",
                           mid_type="num", mid_value=0, mid_color="FFEB84",
                           end_type="max", end_color="00C853"),
        )

    # Signal Detail (ALL raw columns)
    ws = wb.create_sheet("🔬 Signal Detail")
    write_sheet(ws, stock_str_df, "🔬 Signal Detail",
                "🔬  SIGNAL DETAIL  —  All raw columns incl. MST/LST/RS30/ST/EMA/Swing  "
                "|  Reference only — use 📊 Stocks for main view")

    # Apply tab colours
    for ws_obj in wb.worksheets:
        for key, color in TAB_COLORS.items():
            if key in ws_obj.title or ws_obj.title in key:
                ws_obj.sheet_properties.tabColor = color; break
        ws_obj.sheet_view.showGridLines = False

    wb.save(output_path)
    print(f"  💾 Saved: {output_path}")
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS for combining DataFrames
# ─────────────────────────────────────────────────────────────────────────────

def _combine_rotation(sector_rot_df, industry_rot_df):
    """Combine sector rotation and industry rotation into one sheet with a gap row."""
    try:
        if sector_rot_df is None or sector_rot_df.empty:
            return industry_rot_df or pd.DataFrame()
        if industry_rot_df is None or industry_rot_df.empty:
            return sector_rot_df
        # Add a label column if not present
        sr = sector_rot_df.copy()
        ir = industry_rot_df.copy()
        # Align columns
        all_cols = list(dict.fromkeys(list(sr.columns) + list(ir.columns)))
        sr = sr.reindex(columns=all_cols, fill_value="")
        ir = ir.reindex(columns=all_cols, fill_value="")
        gap = pd.DataFrame([{c: ("── INDUSTRY ROTATION ──" if c == (sr.columns[0] if len(sr.columns) > 0 else "Name") else "") for c in all_cols}])
        return pd.concat([sr, gap, ir], ignore_index=True)
    except:
        return sector_rot_df if sector_rot_df is not None else pd.DataFrame()


def _combine_global(country_etf_df, commodity_df):
    """Combine country ETFs and commodities with a gap row."""
    try:
        if country_etf_df is None or country_etf_df.empty:
            return commodity_df or pd.DataFrame()
        if commodity_df is None or commodity_df.empty:
            return country_etf_df
        ce = country_etf_df.copy(); cm = commodity_df.copy()
        all_cols = list(dict.fromkeys(list(ce.columns) + list(cm.columns)))
        ce = ce.reindex(columns=all_cols, fill_value="")
        cm = cm.reindex(columns=all_cols, fill_value="")
        gap_label = ce.columns[1] if len(ce.columns) > 1 else ce.columns[0]
        gap = pd.DataFrame([{c: ("── COMMODITIES (RS vs GLD) ──" if c == gap_label else "") for c in all_cols}])
        return pd.concat([ce, gap, cm], ignore_index=True)
    except:
        return country_etf_df if country_etf_df is not None else pd.DataFrame()


# ─────────────────────────────────────────────────────────────────────────────
#  DASHBOARD SHEET WRITER  v6.0
# ─────────────────────────────────────────────────────────────────────────────

def write_dashboard_sheet(ws, df, market="INDIA"):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 100

    SECTION_BG = "1F3864"; SECTION_FG = "FFFFFF"
    DATA_BG    = "F4F6FB"; ALT_BG     = "FFFFFF"
    GREEN_BG   = "C8E6C9"; GREEN_FG   = "1B5E20"
    RED_BG     = "FFCDD2"; RED_FG     = "B71C1C"
    AMBER_BG   = "FFF9C4"; AMBER_FG   = "5D4037"

    ws.merge_cells("A1:B1")
    t = ws.cell(1, 1, f"📋  FundaTechno Market Analysis  v6.0  [{market}]")
    t.fill = _F("0A1628"); t.font = _fn(bold=True, color="00E5FF", size=14)
    t.alignment = _al("left"); ws.row_dimensions[1].height = 32

    row = 2
    for i, (_, dr) in enumerate(df.iterrows()):
        k = str(dr["Key"]   or "")
        v = str(dr["Value"] or "")
        is_section = k.startswith("──") or k.startswith("══")
        is_blank   = (k == "")

        if is_blank:
            ws.row_dimensions[row].height = 8; row += 1; continue

        if is_section:
            ws.merge_cells(f"A{row}:B{row}")
            c = ws.cell(row, 1, k)
            hbg = "0D47A1" if k.startswith("══") else SECTION_BG
            c.fill = _F(hbg)
            c.font = _fn(bold=True, color="00E5FF" if k.startswith("══") else SECTION_FG, size=11)
            c.alignment = _al("left"); ws.row_dimensions[row].height = 22
        else:
            bg  = DATA_BG if i % 2 == 0 else ALT_BG
            c1  = ws.cell(row, 1, k)
            c2  = ws.cell(row, 2, v)
            c1.fill = _F(bg); c1.font = _fn(bold=True, color="333333", size=10)
            c1.alignment = _al("left")
            c2.fill = _F(bg); c2.font = _fn(size=9, color="222222")
            c2.alignment = _al("left", wrap=True)

            # Color the value cell for key signal counts
            _prime_keys = ("🌟 PRIME BUY", "🌟 Triple", "🌟 RS30", "🌟 Long Momentum", "🌟 Prime")
            _conf_keys  = ("✅ Long Momentum", "✅ Strong RS", "✅ CONFIRMED")
            _rsbuy_keys = ("📈 Swing Entry", "📈 RS Leader", "📈 RS BUY")
            _watch_keys = ("👁",)
            _avoid_keys = ("🔴 RS Breakdown", "🔴 AVOID")

            if any(k.startswith(p) for p in _prime_keys):
                c1.fill = _F("0D3320"); c1.font = _fn(bold=True, color="4ADE80", size=10)
                try:
                    if int(v) > 0: c2.fill = _F("C8E6C9"); c2.font = _fn(bold=True, color="1B5E20")
                except: pass
            elif any(k.startswith(p) for p in _conf_keys):
                c1.fill = _F("1B5E20"); c1.font = _fn(bold=True, color="FFFFFF", size=10)
                try:
                    if int(v) > 0: c2.fill = _F("DCEDC8"); c2.font = _fn(bold=True, color="1B5E20")
                except: pass
            elif any(k.startswith(p) for p in _rsbuy_keys):
                c1.fill = _F("33691E"); c1.font = _fn(bold=True, color="FFFFFF", size=10)
                try:
                    if int(v) > 0: c2.fill = _F("F1F8E9"); c2.font = _fn(bold=True, color="33691E")
                except: pass
            elif any(k.startswith(p) for p in _watch_keys):
                c1.fill = _F("F57F17"); c1.font = _fn(bold=True, color="333333", size=10)
                try:
                    if int(v) > 0: c2.fill = _F(AMBER_BG); c2.font = _fn(color=AMBER_FG)
                except: pass
            elif any(k.startswith(p) for p in _avoid_keys):
                c1.fill = _F("B71C1C"); c1.font = _fn(bold=True, color="FFFFFF", size=10)
                try:
                    if int(v) > 0: c2.fill = _F(RED_BG); c2.font = _fn(color=RED_FG)
                except: pass
            elif k in ("Grade A  ≤3%  (Ideal tight stop)", "Grade B  3-5% (Good)"):
                try:
                    if int(v) > 0: c2.fill = _F(GREEN_BG); c2.font = _fn(bold=True, color=GREEN_FG)
                except: pass

            ws.row_dimensions[row].height = 18
        row += 1

    ws.freeze_panes = "A2"
